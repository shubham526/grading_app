"""
Exam Analytics Tool with Visualizations
Generates comprehensive analytics reports with plots
"""

import os
import sys
import json
import argparse
import numpy as np
from scipy import stats
from datetime import datetime
import csv
from collections import defaultdict
import matplotlib.pyplot as plt
import matplotlib

matplotlib.use('Agg')  # Use non-interactive backend
import seaborn as sns

# Set style for better-looking plots
sns.set_style("whitegrid")
plt.rcParams['figure.figsize'] = (10, 6)
plt.rcParams['font.size'] = 10


def load_graded_exams(folder_path):
    """Load all JSON graded exam files from a folder."""
    exams = []

    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist!")
        return exams

    json_files = [f for f in os.listdir(folder_path) if f.endswith('.json')]

    if not json_files:
        print(f"Warning: No JSON files found in '{folder_path}'")
        return exams

    print(f"Found {len(json_files)} JSON files")

    for filename in json_files:
        filepath = os.path.join(folder_path, filename)
        try:
            with open(filepath, 'r') as f:
                data = json.load(f)
                exams.append(data)
                print(f"  ✓ Loaded: {filename}")
        except Exception as e:
            print(f"  ✗ Error loading {filename}: {e}")

    return exams


def extract_analytics_data(exams):
    """Extract data needed for analytics from graded exam JSON files."""
    question_data = defaultdict(lambda: {'scores': [], 'max_points': 0})
    overall_scores = []
    student_info = []
    assignment_name = None

    # Also collect full question-by-student matrix for advanced analysis
    student_question_matrix = []

    for exam in exams:
        if assignment_name is None:
            assignment_name = exam.get('assignment_name', 'Unknown Assignment')

        student_name = exam.get('student_name', 'Unknown Student')
        total_awarded = exam.get('total_awarded', 0)
        total_possible = exam.get('total_possible', 0)
        percentage = exam.get('percentage', 0)

        overall_scores.append(percentage)

        if percentage >= 90:
            grade = 'A'
        elif percentage >= 80:
            grade = 'B'
        elif percentage >= 70:
            grade = 'C'
        elif percentage >= 60:
            grade = 'D'
        else:
            grade = 'F'

        student_info.append({
            'name': student_name,
            'score': total_awarded,
            'max_score': total_possible,
            'percentage': percentage,
            'grade': grade
        })

        question_summary = exam.get('question_summary', [])
        student_row = {}

        for q in question_summary:
            if q.get('counted', False):
                q_num = q.get('question')
                awarded = q.get('awarded', 0)
                possible = q.get('possible', 0)

                question_data[q_num]['scores'].append(awarded)
                question_data[q_num]['max_points'] = possible
                student_row[q_num] = awarded

        student_question_matrix.append(student_row)

    question_data = dict(question_data)

    return question_data, overall_scores, student_info, assignment_name, student_question_matrix


def calculate_grade_distribution(overall_scores):
    """Calculate grade distribution."""
    grades = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}

    for score in overall_scores:
        if score >= 90:
            grades['A'] += 1
        elif score >= 80:
            grades['B'] += 1
        elif score >= 70:
            grades['C'] += 1
        elif score >= 60:
            grades['D'] += 1
        else:
            grades['F'] += 1

    return grades


def calculate_question_difficulty(question_data):
    """Calculate difficulty index for each question."""
    difficulty = {}

    for q_num, data in question_data.items():
        if data['max_points'] > 0:
            avg_score = np.mean(data['scores'])
            difficulty[q_num] = (avg_score / data['max_points']) * 100
        else:
            difficulty[q_num] = 0

    return difficulty


def calculate_discrimination_index(question_data, overall_scores):
    """Calculate discrimination index for each question."""
    discrimination = {}

    n = len(overall_scores)
    cutoff = max(1, int(n * 0.27))  # At least 1 student in each group

    sorted_indices = np.argsort(overall_scores)
    lower_group_indices = sorted_indices[:cutoff]
    upper_group_indices = sorted_indices[-cutoff:]

    for q_num, data in question_data.items():
        scores = np.array(data['scores'])
        max_points = data['max_points']

        # Filter indices to only those valid for this question's scores array
        valid_lower = lower_group_indices[lower_group_indices < len(scores)]
        valid_upper = upper_group_indices[upper_group_indices < len(scores)]

        # Need at least 1 student in each group for meaningful discrimination
        if max_points > 0 and len(valid_lower) >= 1 and len(valid_upper) >= 1:
            upper_avg = np.mean(scores[valid_upper])
            lower_avg = np.mean(scores[valid_lower])

            discrimination[q_num] = (upper_avg - lower_avg) / max_points
        else:
            discrimination[q_num] = 0

    return discrimination


def get_discrimination_quality(value):
    """Get quality rating for discrimination index."""
    if value >= 0.40:
        return "Excellent"
    elif value >= 0.30:
        return "Good"
    elif value >= 0.20:
        return "Fair"
    else:
        return "Poor"


def calculate_cronbach_alpha(question_data):
    """Calculate Cronbach's Alpha for exam reliability."""
    questions = sorted(question_data.keys())

    if len(questions) < 2:
        return None

    n_students = len(question_data[questions[0]]['scores'])

    score_matrix = []
    for q in questions:
        scores = question_data[q]['scores']
        if len(scores) == n_students:
            score_matrix.append(scores)

    if len(score_matrix) < 2:
        return None

    score_matrix = np.array(score_matrix).T
    n_items = score_matrix.shape[1]

    item_variances = np.var(score_matrix, axis=0, ddof=1)
    total_scores = np.sum(score_matrix, axis=1)
    total_variance = np.var(total_scores, ddof=1)

    if total_variance == 0:
        return None

    alpha = (n_items / (n_items - 1)) * (1 - np.sum(item_variances) / total_variance)

    return float(alpha)


def get_reliability_interpretation(alpha):
    """Get interpretation of Cronbach's alpha value."""
    if alpha is None:
        return "Cannot calculate"
    elif alpha >= 0.9:
        return "Excellent reliability"
    elif alpha >= 0.8:
        return "Good reliability"
    elif alpha >= 0.7:
        return "Acceptable reliability"
    elif alpha >= 0.6:
        return "Questionable reliability"
    else:
        return "Poor reliability"


# ============================================================================
# PLOTTING FUNCTIONS
# ============================================================================

def plot_score_distribution(overall_scores, output_dir, assignment_name):
    """Create histogram with normal curve overlay."""
    plt.figure(figsize=(12, 6))

    # Histogram
    n, bins, patches = plt.hist(overall_scores, bins=15, alpha=0.7,
                                color='steelblue', edgecolor='black',
                                density=True, label='Actual Distribution')

    # Fit normal curve
    mu = np.mean(overall_scores)
    sigma = np.std(overall_scores)
    x = np.linspace(0, 100, 100)
    y = stats.norm.pdf(x, mu, sigma)
    plt.plot(x, y, 'r-', linewidth=2, label=f'Normal (μ={mu:.1f}, σ={sigma:.1f})')

    plt.xlabel('Score (%)', fontsize=12, fontweight='bold')
    plt.ylabel('Density', fontsize=12, fontweight='bold')
    plt.title(f'Score Distribution\n{assignment_name}', fontsize=14, fontweight='bold')
    plt.legend()
    plt.grid(True, alpha=0.3)

    output_path = os.path.join(output_dir, 'score_distribution.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  ✓ Created: score_distribution.png")


def plot_box_plot(overall_scores, output_dir, assignment_name):
    """Create box plot showing score distribution."""
    plt.figure(figsize=(10, 6))

    box = plt.boxplot([overall_scores], vert=True, patch_artist=True,
                      labels=['Exam Scores'],
                      widths=0.5)

    # Color the box
    box['boxes'][0].set_facecolor('lightblue')
    box['boxes'][0].set_edgecolor('black')
    box['boxes'][0].set_linewidth(2)

    # Add statistical annotations
    q1 = np.percentile(overall_scores, 25)
    median = np.median(overall_scores)
    q3 = np.percentile(overall_scores, 75)

    plt.text(1.2, q1, f'Q1: {q1:.1f}%', fontsize=10, va='center')
    plt.text(1.2, median, f'Median: {median:.1f}%', fontsize=10, va='center', fontweight='bold')
    plt.text(1.2, q3, f'Q3: {q3:.1f}%', fontsize=10, va='center')

    plt.ylabel('Score (%)', fontsize=12, fontweight='bold')
    plt.title(f'Score Distribution Box Plot\n{assignment_name}', fontsize=14, fontweight='bold')
    plt.grid(True, alpha=0.3, axis='y')

    output_path = os.path.join(output_dir, 'box_plot.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  ✓ Created: box_plot.png")


def plot_grade_distribution(grades, output_dir, assignment_name):
    """Create bar chart of grade distribution."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

    # Bar chart
    grade_labels = ['A', 'B', 'C', 'D', 'F']
    counts = [grades[g] for g in grade_labels]
    colors = ['#2ecc71', '#3498db', '#f39c12', '#e67e22', '#e74c3c']

    bars = ax1.bar(grade_labels, counts, color=colors, edgecolor='black', linewidth=1.5)
    ax1.set_xlabel('Grade', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Number of Students', fontsize=12, fontweight='bold')
    ax1.set_title('Grade Distribution', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3, axis='y')

    # Add count labels on bars
    for bar in bars:
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width() / 2., height,
                 f'{int(height)}',
                 ha='center', va='bottom', fontweight='bold')

    # Pie chart
    total = sum(counts)
    if total > 0:
        percentages = [(c / total) * 100 for c in counts]
        # Only show non-zero grades
        filtered_labels = [grade_labels[i] for i in range(len(counts)) if counts[i] > 0]
        filtered_counts = [counts[i] for i in range(len(counts)) if counts[i] > 0]
        filtered_colors = [colors[i] for i in range(len(counts)) if counts[i] > 0]

        wedges, texts, autotexts = ax2.pie(filtered_counts, labels=filtered_labels,
                                           colors=filtered_colors, autopct='%1.1f%%',
                                           startangle=90, textprops={'fontweight': 'bold'})
        ax2.set_title('Grade Percentages', fontsize=14, fontweight='bold')

    plt.suptitle(f'{assignment_name}', fontsize=16, fontweight='bold', y=1.02)

    output_path = os.path.join(output_dir, 'grade_distribution.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  ✓ Created: grade_distribution.png")


def plot_question_difficulty(difficulty, output_dir, assignment_name):
    """Create bar chart showing difficulty of each question."""
    plt.figure(figsize=(12, 6))

    questions = sorted(difficulty.keys())
    diff_values = [difficulty[q] for q in questions]

    # Color bars by difficulty
    colors = []
    for d in diff_values:
        if d >= 80:
            colors.append('#2ecc71')  # Easy - Green
        elif d >= 60:
            colors.append('#f39c12')  # Medium - Orange
        else:
            colors.append('#e74c3c')  # Hard - Red

    bars = plt.bar(range(len(questions)), diff_values, color=colors,
                   edgecolor='black', linewidth=1.5)

    plt.xlabel('Question', fontsize=12, fontweight='bold')
    plt.ylabel('Difficulty (%)', fontsize=12, fontweight='bold')
    plt.title(f'Question Difficulty Analysis\n{assignment_name}',
              fontsize=14, fontweight='bold')
    plt.xticks(range(len(questions)), [f"Q{q}" for q in questions])
    plt.axhline(y=80, color='g', linestyle='--', alpha=0.5, label='Easy (≥80%)')
    plt.axhline(y=60, color='orange', linestyle='--', alpha=0.5, label='Medium (60-80%)')
    plt.axhline(y=40, color='r', linestyle='--', alpha=0.5, label='Hard (<60%)')
    plt.legend()
    plt.grid(True, alpha=0.3, axis='y')

    # Add value labels
    for i, bar in enumerate(bars):
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width() / 2., height,
                 f'{diff_values[i]:.1f}%',
                 ha='center', va='bottom', fontsize=9, fontweight='bold')

    output_path = os.path.join(output_dir, 'question_difficulty.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  ✓ Created: question_difficulty.png")


def plot_discrimination_index(discrimination, output_dir, assignment_name):
    """Create bar chart showing discrimination index for each question."""
    plt.figure(figsize=(12, 6))

    questions = sorted(discrimination.keys())
    disc_values = [discrimination[q] for q in questions]

    # Color bars by quality
    colors = []
    for d in disc_values:
        if d >= 0.40:
            colors.append('#2ecc71')  # Excellent - Green
        elif d >= 0.30:
            colors.append('#3498db')  # Good - Blue
        elif d >= 0.20:
            colors.append('#f39c12')  # Fair - Orange
        else:
            colors.append('#e74c3c')  # Poor - Red

    bars = plt.bar(range(len(questions)), disc_values, color=colors,
                   edgecolor='black', linewidth=1.5)

    plt.xlabel('Question', fontsize=12, fontweight='bold')
    plt.ylabel('Discrimination Index', fontsize=12, fontweight='bold')
    plt.title(f'Question Discrimination Analysis\n{assignment_name}',
              fontsize=14, fontweight='bold')
    plt.xticks(range(len(questions)), [f"Q{q}" for q in questions])
    plt.axhline(y=0.40, color='g', linestyle='--', alpha=0.5, label='Excellent (≥0.40)')
    plt.axhline(y=0.30, color='b', linestyle='--', alpha=0.5, label='Good (≥0.30)')
    plt.axhline(y=0.20, color='orange', linestyle='--', alpha=0.5, label='Fair (≥0.20)')
    plt.legend()
    plt.grid(True, alpha=0.3, axis='y')

    # Add value labels
    for i, bar in enumerate(bars):
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width() / 2., height,
                 f'{disc_values[i]:.3f}',
                 ha='center', va='bottom', fontsize=9, fontweight='bold')

    output_path = os.path.join(output_dir, 'discrimination_index.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  ✓ Created: discrimination_index.png")


def plot_difficulty_vs_discrimination(difficulty, discrimination, output_dir, assignment_name):
    """Scatter plot showing relationship between difficulty and discrimination."""
    plt.figure(figsize=(10, 8))

    questions = sorted(difficulty.keys())
    diff_values = [difficulty[q] for q in questions]
    disc_values = [discrimination[q] for q in questions]

    # Color points by quality
    colors = []
    for d in disc_values:
        if d >= 0.40:
            colors.append('#2ecc71')
        elif d >= 0.30:
            colors.append('#3498db')
        elif d >= 0.20:
            colors.append('#f39c12')
        else:
            colors.append('#e74c3c')

    plt.scatter(diff_values, disc_values, c=colors, s=200, alpha=0.7,
                edgecolors='black', linewidth=2)

    # Add question labels
    for i, q in enumerate(questions):
        plt.annotate(f'Q{q}', (diff_values[i], disc_values[i]),
                     fontsize=10, fontweight='bold', ha='center', va='center')

    # Add quadrant lines
    plt.axhline(y=0.30, color='gray', linestyle='--', alpha=0.5)
    plt.axvline(x=70, color='gray', linestyle='--', alpha=0.5)

    # Add quadrant labels
    plt.text(85, 0.45, 'Easy &\nGood Disc.', ha='center', fontsize=10,
             bbox=dict(boxstyle='round', facecolor='lightgreen', alpha=0.5))
    plt.text(85, 0.10, 'Easy &\nPoor Disc.', ha='center', fontsize=10,
             bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.5))
    plt.text(40, 0.45, 'Hard &\nGood Disc.', ha='center', fontsize=10,
             bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.5))
    plt.text(40, 0.10, 'Hard &\nPoor Disc.', ha='center', fontsize=10,
             bbox=dict(boxstyle='round', facecolor='lightcoral', alpha=0.5))

    plt.xlabel('Difficulty (% Correct)', fontsize=12, fontweight='bold')
    plt.ylabel('Discrimination Index', fontsize=12, fontweight='bold')
    plt.title(f'Question Quality Analysis\n{assignment_name}',
              fontsize=14, fontweight='bold')
    plt.grid(True, alpha=0.3)

    output_path = os.path.join(output_dir, 'difficulty_vs_discrimination.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  ✓ Created: difficulty_vs_discrimination.png")


def plot_student_performance_heatmap(student_info, question_data, student_question_matrix,
                                     output_dir, assignment_name):
    """Create heatmap of student performance on each question."""
    if not student_question_matrix:
        return

    # Build matrix
    questions = sorted(question_data.keys())
    students = [s['name'] for s in student_info]

    # Create percentage matrix
    matrix = []
    for student_row in student_question_matrix:
        row = []
        for q in questions:
            if q in student_row:
                max_points = question_data[q]['max_points']
                pct = (student_row[q] / max_points * 100) if max_points > 0 else 0
                row.append(pct)
            else:
                row.append(0)
        matrix.append(row)

    matrix = np.array(matrix)

    # Sort students by overall performance
    overall_scores = [s['percentage'] for s in student_info]
    sorted_indices = np.argsort(overall_scores)[::-1]  # Descending
    matrix = matrix[sorted_indices]
    students = [students[i] for i in sorted_indices]

    # Limit to top 30 students for readability
    if len(students) > 30:
        matrix = matrix[:30]
        students = students[:30]

    plt.figure(figsize=(max(12, len(questions) * 0.8), max(8, len(students) * 0.3)))

    sns.heatmap(matrix, annot=True, fmt='.0f', cmap='RdYlGn',
                vmin=0, vmax=100, cbar_kws={'label': 'Score (%)'},
                xticklabels=[f'Q{q}' for q in questions],
                yticklabels=students, linewidths=0.5)

    plt.xlabel('Question', fontsize=12, fontweight='bold')
    plt.ylabel('Student (sorted by overall score)', fontsize=12, fontweight='bold')
    plt.title(f'Student Performance Heatmap\n{assignment_name}',
              fontsize=14, fontweight='bold')
    plt.tight_layout()

    output_path = os.path.join(output_dir, 'performance_heatmap.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  ✓ Created: performance_heatmap.png")


def plot_student_ranking(student_info, output_dir, assignment_name):
    """Create bar chart showing student rankings."""
    # Sort students by score
    sorted_students = sorted(student_info, key=lambda x: x['percentage'], reverse=True)

    # Limit to top 20 for readability
    if len(sorted_students) > 20:
        sorted_students = sorted_students[:20]
        title_suffix = " (Top 20)"
    else:
        title_suffix = ""

    names = [s['name'] for s in sorted_students]
    scores = [s['percentage'] for s in sorted_students]
    grades = [s['grade'] for s in sorted_students]

    # Color by grade
    grade_colors = {'A': '#2ecc71', 'B': '#3498db', 'C': '#f39c12',
                    'D': '#e67e22', 'F': '#e74c3c'}
    colors = [grade_colors[g] for g in grades]

    plt.figure(figsize=(12, max(8, len(names) * 0.4)))

    bars = plt.barh(range(len(names)), scores, color=colors,
                    edgecolor='black', linewidth=1.5)

    plt.yticks(range(len(names)), names)
    plt.xlabel('Score (%)', fontsize=12, fontweight='bold')
    plt.ylabel('Student', fontsize=12, fontweight='bold')
    plt.title(f'Student Rankings{title_suffix}\n{assignment_name}',
              fontsize=14, fontweight='bold')
    plt.grid(True, alpha=0.3, axis='x')

    # Add score and grade labels
    for i, (bar, score, grade) in enumerate(zip(bars, scores, grades)):
        width = bar.get_width()
        plt.text(width + 1, bar.get_y() + bar.get_height() / 2.,
                 f'{score:.1f}% ({grade})',
                 ha='left', va='center', fontsize=9, fontweight='bold')

    plt.xlim(0, 110)
    plt.tight_layout()

    output_path = os.path.join(output_dir, 'student_ranking.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  ✓ Created: student_ranking.png")


def generate_all_plots(analytics_data, question_data, student_question_matrix, output_dir):
    """Generate all visualization plots."""
    print("\nGenerating plots...")

    assignment_name = analytics_data['assignment_name']
    overall_scores = [s['percentage'] for s in analytics_data['student_scores']]
    grades = analytics_data['grade_distribution']

    difficulty = {q: analytics_data['question_stats'][q]['difficulty']
                  for q in analytics_data['question_stats']}
    discrimination = {q: analytics_data['question_stats'][q]['discrimination']
                      for q in analytics_data['question_stats']}

    # Generate each plot
    plot_score_distribution(overall_scores, output_dir, assignment_name)
    plot_box_plot(overall_scores, output_dir, assignment_name)
    plot_grade_distribution(grades, output_dir, assignment_name)
    plot_question_difficulty(difficulty, output_dir, assignment_name)
    plot_discrimination_index(discrimination, output_dir, assignment_name)
    plot_difficulty_vs_discrimination(difficulty, discrimination, output_dir, assignment_name)
    plot_student_performance_heatmap(analytics_data['student_scores'], question_data,
                                     student_question_matrix, output_dir, assignment_name)
    plot_student_ranking(analytics_data['student_scores'], output_dir, assignment_name)

    print(f"\n✓ All plots saved to: {output_dir}/")


def generate_analytics_report(exams):
    """Generate complete analytics report from exam data."""
    if not exams:
        print("No exam data to analyze!")
        return None, None, None

    question_data, overall_scores, student_info, assignment_name, student_question_matrix = extract_analytics_data(
        exams)

    if not overall_scores:
        print("No student scores found!")
        return None, None, None

    print(f"\nAnalyzing {len(overall_scores)} students...")

    grades = calculate_grade_distribution(overall_scores)
    difficulty = calculate_question_difficulty(question_data)
    discrimination = calculate_discrimination_index(question_data, overall_scores)
    alpha = calculate_cronbach_alpha(question_data)

    question_stats = {}
    for q_num in sorted(question_data.keys()):
        scores = question_data[q_num]['scores']
        max_pts = question_data[q_num]['max_points']

        question_stats[q_num] = {
            'mean_score': float(np.mean(scores)),
            'mean_percentage': difficulty.get(q_num, 0),
            'difficulty': difficulty.get(q_num, 0),
            'discrimination': discrimination.get(q_num, 0),
            'discrimination_quality': get_discrimination_quality(discrimination.get(q_num, 0)),
            'max_points': max_pts,
            'std_dev': float(np.std(scores))
        }

    analytics_data = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'assignment_name': assignment_name,
        'num_students': len(overall_scores),
        'overall_stats': {
            'mean': float(np.mean(overall_scores)),
            'median': float(np.median(overall_scores)),
            'std_dev': float(np.std(overall_scores)),
            'min': float(np.min(overall_scores)),
            'max': float(np.max(overall_scores))
        },
        'grade_distribution': grades,
        'question_stats': question_stats,
        'cronbach_alpha': f"{alpha:.3f}" if alpha else "N/A",
        'reliability_interpretation': get_reliability_interpretation(alpha),
        'student_scores': student_info
    }

    return analytics_data, question_data, student_question_matrix


def print_analytics_summary(analytics_data):
    """Print a formatted summary of analytics to console."""
    if not analytics_data:
        return

    print("\n" + "=" * 70)
    print(f"ANALYTICS REPORT: {analytics_data['assignment_name']}")
    print("=" * 70)

    print(f"\nGenerated: {analytics_data['timestamp']}")
    print(f"Number of Students: {analytics_data['num_students']}")

    print("\n" + "-" * 70)
    print("OVERALL STATISTICS")
    print("-" * 70)
    stats = analytics_data['overall_stats']
    print(f"  Mean Score:       {stats['mean']:.2f}%")
    print(f"  Median Score:     {stats['median']:.2f}%")
    print(f"  Std Deviation:    {stats['std_dev']:.2f}")
    print(f"  Minimum Score:    {stats['min']:.2f}%")
    print(f"  Maximum Score:    {stats['max']:.2f}%")

    print("\n" + "-" * 70)
    print("GRADE DISTRIBUTION")
    print("-" * 70)
    grades = analytics_data['grade_distribution']
    total = sum(grades.values())
    for grade in ['A', 'B', 'C', 'D', 'F']:
        count = grades[grade]
        pct = (count / total * 100) if total > 0 else 0
        bar = "█" * int(pct / 2)
        print(f"  {grade}: {count:2d} students ({pct:5.1f}%) {bar}")

    print("\n" + "-" * 70)
    print("QUESTION ANALYSIS")
    print("-" * 70)
    print(f"{'Question':<12} {'Mean %':<10} {'Difficulty':<12} {'Discrimination':<15} {'Quality'}")
    print("-" * 70)

    for q_num in sorted(analytics_data['question_stats'].keys()):
        stats = analytics_data['question_stats'][q_num]
        print(f"Question {q_num:<4} {stats['mean_percentage']:>6.1f}%    "
              f"{stats['difficulty']:>6.1f}%       "
              f"{stats['discrimination']:>6.3f}          "
              f"{stats['discrimination_quality']}")

    print("\n" + "-" * 70)
    print("RELIABILITY ANALYSIS")
    print("-" * 70)
    print(f"  Cronbach's Alpha: {analytics_data['cronbach_alpha']}")
    print(f"  Interpretation:   {analytics_data['reliability_interpretation']}")

    print("\n" + "=" * 70)


def export_to_csv(analytics_data, output_path):
    """Export analytics to CSV file."""
    with open(output_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)

        writer.writerow(['Analytics Report'])
        writer.writerow(['Generated:', analytics_data['timestamp']])
        writer.writerow(['Assignment:', analytics_data['assignment_name']])
        writer.writerow(['Number of Students:', analytics_data['num_students']])
        writer.writerow([])

        writer.writerow(['OVERALL STATISTICS'])
        writer.writerow(['Metric', 'Value'])
        overall = analytics_data['overall_stats']
        writer.writerow(['Mean Score', f"{overall['mean']:.2f}%"])
        writer.writerow(['Median Score', f"{overall['median']:.2f}%"])
        writer.writerow(['Standard Deviation', f"{overall['std_dev']:.2f}"])
        writer.writerow(['Minimum Score', f"{overall['min']:.2f}%"])
        writer.writerow(['Maximum Score', f"{overall['max']:.2f}%"])
        writer.writerow([])

        writer.writerow(['GRADE DISTRIBUTION'])
        writer.writerow(['Grade', 'Count', 'Percentage'])
        grades = analytics_data['grade_distribution']
        total = sum(grades.values())
        for grade in ['A', 'B', 'C', 'D', 'F']:
            count = grades[grade]
            pct = (count / total * 100) if total > 0 else 0
            writer.writerow([grade, count, f"{pct:.1f}%"])
        writer.writerow([])

        writer.writerow(['QUESTION ANALYSIS'])
        writer.writerow(['Question', 'Mean %', 'Difficulty', 'Discrimination', 'Quality'])

        for q_num in sorted(analytics_data['question_stats'].keys()):
            stats = analytics_data['question_stats'][q_num]
            writer.writerow([
                f"Question {q_num}",
                f"{stats['mean_percentage']:.1f}%",
                f"{stats['difficulty']:.1f}%",
                f"{stats['discrimination']:.3f}",
                stats['discrimination_quality']
            ])
        writer.writerow([])

        writer.writerow(['RELIABILITY ANALYSIS'])
        writer.writerow(['Cronbach\'s Alpha', analytics_data['cronbach_alpha']])
        writer.writerow(['Interpretation', analytics_data['reliability_interpretation']])
        writer.writerow([])

        writer.writerow(['STUDENT SCORES'])
        writer.writerow(['Student', 'Score', 'Percentage', 'Grade'])

        for student in sorted(analytics_data['student_scores'], key=lambda x: x['percentage'], reverse=True):
            writer.writerow([
                student['name'],
                f"{student['score']:.1f} / {student['max_score']:.1f}",
                f"{student['percentage']:.1f}%",
                student['grade']
            ])


def export_to_excel(analytics_data, output_path):
    """Export analytics to Excel file with multiple sheets."""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Summary
            summary_data = {
                'Metric': [
                    'Assignment Name',
                    'Number of Students',
                    'Mean Score (%)',
                    'Median Score (%)',
                    'Standard Deviation',
                    'Minimum Score (%)',
                    'Maximum Score (%)',
                    'Cronbach\'s Alpha',
                    'Reliability'
                ],
                'Value': [
                    analytics_data['assignment_name'],
                    analytics_data['num_students'],
                    f"{analytics_data['overall_stats']['mean']:.2f}",
                    f"{analytics_data['overall_stats']['median']:.2f}",
                    f"{analytics_data['overall_stats']['std_dev']:.2f}",
                    f"{analytics_data['overall_stats']['min']:.2f}",
                    f"{analytics_data['overall_stats']['max']:.2f}",
                    analytics_data['cronbach_alpha'],
                    analytics_data['reliability_interpretation']
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)

            # Sheet 2: Grade Distribution
            grades = analytics_data['grade_distribution']
            total = sum(grades.values())
            grade_data = {
                'Grade': ['A', 'B', 'C', 'D', 'F'],
                'Count': [grades[g] for g in ['A', 'B', 'C', 'D', 'F']],
                'Percentage': [f"{(grades[g] / total * 100):.1f}%" if total > 0 else "0.0%"
                               for g in ['A', 'B', 'C', 'D', 'F']]
            }
            df_grades = pd.DataFrame(grade_data)
            df_grades.to_excel(writer, sheet_name='Grade Distribution', index=False)

            # Sheet 3: Question Analysis
            question_data = []
            for q_num in sorted(analytics_data['question_stats'].keys()):
                stats = analytics_data['question_stats'][q_num]
                question_data.append({
                    'Question': f"Question {q_num}",
                    'Mean %': f"{stats['mean_percentage']:.1f}",
                    'Difficulty %': f"{stats['difficulty']:.1f}",
                    'Discrimination': f"{stats['discrimination']:.3f}",
                    'Quality': stats['discrimination_quality']
                })
            df_questions = pd.DataFrame(question_data)
            df_questions.to_excel(writer, sheet_name='Question Analysis', index=False)

            # Sheet 4: Student Scores
            student_data = []
            for student in sorted(analytics_data['student_scores'], key=lambda x: x['percentage'], reverse=True):
                student_data.append({
                    'Student': student['name'],
                    'Score': f"{student['score']:.1f}",
                    'Max Score': f"{student['max_score']:.1f}",
                    'Percentage': f"{student['percentage']:.1f}",
                    'Grade': student['grade']
                })
            df_students = pd.DataFrame(student_data)
            df_students.to_excel(writer, sheet_name='Student Scores', index=False)

        # Format the Excel file
        wb = load_workbook(output_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Header formatting
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        print(f"✓ Excel report exported to: {output_path}")

    except ImportError:
        print("Warning: pandas or openpyxl not available. Falling back to CSV.")
        csv_path = output_path.replace('.xlsx', '.csv')
        export_to_csv(analytics_data, csv_path)
        print(f"✓ CSV report exported to: {csv_path}")


def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description='Generate analytics reports with visualizations from graded exam JSON files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage with plots
  python exam_analytics.py -i ./graded_exams/

  # Specify custom output and generate plots
  python exam_analytics.py -i ./graded_exams/ -o report.csv --plots

  # Generate Excel report with plots in separate folder
  python exam_analytics.py -i ./graded_exams/ -o report.xlsx --format excel --plots --plot-dir ./plots/

  # Skip plots generation
  python exam_analytics.py -i ./graded_exams/ --no-plots
        """
    )

    parser.add_argument(
        '-i', '--input',
        required=True,
        help='Input folder containing graded exam JSON files',
        metavar='FOLDER'
    )

    parser.add_argument(
        '-o', '--output',
        help='Output file path (default: analytics_report.csv in input folder)',
        metavar='FILE'
    )

    parser.add_argument(
        '--format',
        choices=['csv', 'excel'],
        default='csv',
        help='Output format: csv or excel (default: csv)'
    )

    parser.add_argument(
        '--plots',
        action='store_true',
        default=True,
        help='Generate visualization plots (default: True)'
    )

    parser.add_argument(
        '--no-plots',
        action='store_true',
        help='Skip generating plots'
    )

    parser.add_argument(
        '--plot-dir',
        help='Directory to save plots (default: plots/ in input folder)',
        metavar='DIR'
    )

    parser.add_argument(
        '--no-console',
        action='store_true',
        help='Suppress console output (only save to file)'
    )

    return parser.parse_args()


def main():
    """Main function to run analytics from command line."""
    args = parse_arguments()

    # Validate input folder
    input_folder = args.input
    if not os.path.exists(input_folder):
        print(f"Error: Input folder '{input_folder}' does not exist!")
        sys.exit(1)

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        if args.format == 'excel':
            output_path = os.path.join(input_folder, 'analytics_report.xlsx')
        else:
            output_path = os.path.join(input_folder, 'analytics_report.csv')

    # Ensure correct file extension
    if args.format == 'excel' and not output_path.endswith('.xlsx'):
        output_path = output_path.rsplit('.', 1)[0] + '.xlsx'
    elif args.format == 'csv' and not output_path.endswith('.csv'):
        output_path = output_path.rsplit('.', 1)[0] + '.csv'

    # Determine plot directory
    if args.plot_dir:
        plot_dir = args.plot_dir
    else:
        plot_dir = os.path.join(input_folder, 'plots')

    # Create plot directory if needed
    generate_plots = args.plots and not args.no_plots
    if generate_plots:
        os.makedirs(plot_dir, exist_ok=True)

    # Print header
    if not args.no_console:
        print(f"\n{'=' * 70}")
        print("EXAM ANALYTICS TOOL WITH VISUALIZATIONS")
        print(f"{'=' * 70}")
        print(f"\nInput folder: {input_folder}")
        print(f"Output file:  {output_path}")
        print(f"Format:       {args.format.upper()}")
        if generate_plots:
            print(f"Plot folder:  {plot_dir}")

    # Load exam files
    exams = load_graded_exams(input_folder)

    if not exams:
        print("\nNo exams loaded. Exiting.")
        sys.exit(1)

    # Generate analytics
    analytics_data, question_data, student_question_matrix = generate_analytics_report(exams)

    if not analytics_data:
        print("\nFailed to generate analytics. Exiting.")
        sys.exit(1)

    # Print summary to console (unless suppressed)
    if not args.no_console:
        print_analytics_summary(analytics_data)

    # Export to file
    if args.format == 'excel':
        export_to_excel(analytics_data, output_path)
    else:
        export_to_csv(analytics_data, output_path)
        print(f"\n✓ CSV report exported to: {output_path}")

    # Generate plots
    if generate_plots:
        generate_all_plots(analytics_data, question_data, student_question_matrix, plot_dir)

    if not args.no_console:
        print("\n" + "=" * 70)
        print("ANALYSIS COMPLETE!")
        print("=" * 70)
        if generate_plots:
            print(f"\n📊 Generated {8} visualization plots in: {plot_dir}/")
        print(f"📄 Report saved to: {output_path}")


if __name__ == "__main__":
    main()