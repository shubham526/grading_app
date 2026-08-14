"""Master ABET evidence generation and export for v2.2.1.

The backend produces one auditable row per student/assignment/rubric criterion,
composes those rows across semester configs, and exports the normalized evidence
to CSV, JSON, and XLSX without altering any grading or ABET score calculation.

The implementation follows the application's existing persistence semantics:

* Saved assessment criteria are the snapshot used by current ABET scoring and
  therefore remain primary for final points, selected/counted state, and any
  outcome metadata already persisted with the grade.
* The rubric is joined by stable criterion ID and fills metadata that is absent
  from older assessment files.  Legacy title matching is used only when a
  stable ID is missing on one side.
* Evidence-policy filtering matches ``src.tools.abet_scoring``.  In particular,
  legacy assessment criteria that predate selected/counted flags retain the
  scoring engine's historical default of being included.  The exported raw
  flag remains ``None`` so missing data is still transparent.
* v2.2 submission metadata is optional.  Missing submission ingestion never
  prevents evidence-row generation.
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from src.tools.abet_scoring import (
    DEFAULT_POLICY,
    POLICY_ALL,
    POLICY_COUNTED,
    POLICY_SELECTED,
)


VALID_EVIDENCE_POLICIES = frozenset({
    POLICY_COUNTED,
    POLICY_SELECTED,
    POLICY_ALL,
})

# Stable internal schema.  Export-format-specific display names are deliberately
# deferred to the later CSV/XLSX/JSON commit.
MASTER_EVIDENCE_FIELDS: Tuple[str, ...] = (
    "semester",
    "course_code",
    "course_name",
    "section",
    "assignment_id",
    "assignment_title",
    "assignment_type",
    "assignment_date",
    "student_id",
    "student_name",
    "question_id",
    "criterion_id",
    "criterion_title",
    "criterion_description",
    "points_awarded",
    "points_possible",
    "percentage",
    "selected",
    "counted",
    "evidence_policy",
    "course_outcomes",
    "program_outcomes",
    "abet_outcomes",
    "assessment_tags",
    "performance_band",
    "meets_target",
    "submission_source",
    "submission_file_latex",
    "submission_file_pdf",
    "submission_hash_latex",
    "submission_hash_pdf",
    "notes",
)


@dataclass
class MasterEvidenceBuildResult:
    """Rows plus non-fatal data-quality warnings collected while building."""

    rows: List[dict] = field(default_factory=list)
    warnings: List[dict] = field(default_factory=list)


def _warning(
    code: str,
    message: str,
    *,
    assessment_file: str = "",
    student_id: str = "",
    criterion_id: str = "",
) -> dict:
    return {
        "code": str(code),
        "message": str(message),
        "assessment_file": str(assessment_file or ""),
        "student_id": str(student_id or ""),
        "criterion_id": str(criterion_id or ""),
    }


def _nonblank(*values: Any, default: Any = "") -> Any:
    """Return the first value that is not None/blank-string."""
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return default


def _string_list(value: Any) -> List[str]:
    """Normalize a list-like metadata field without inventing values."""
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        return [text] if text else []
    if isinstance(value, (list, tuple, set)):
        result: List[str] = []
        for item in value:
            text = str(item).strip() if item is not None else ""
            if text and text not in result:
                result.append(text)
        return result
    text = str(value).strip()
    return [text] if text else []


def _number_or_none(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _percentage(points_awarded: Any, points_possible: Any) -> Optional[float]:
    awarded = _number_or_none(points_awarded)
    possible = _number_or_none(points_possible)
    if awarded is None or possible is None or possible == 0:
        return None
    return awarded / possible * 100.0


def _raw_bool(criterion: Optional[Mapping[str, Any]], key: str) -> Optional[bool]:
    """Return the persisted flag, preserving missingness as None."""
    if not criterion or key not in criterion or criterion.get(key) is None:
        return None
    return bool(criterion.get(key))


def _criterion_included_for_policy(
    assessment_criterion: Optional[Mapping[str, Any]],
    policy: str,
) -> bool:
    """Match current ABET scoring inclusion semantics.

    ``abet_scoring._criterion_included`` defaults missing selected/counted flags
    to True for legacy criteria.  A rubric-only criterion, however, has no saved
    student evidence at all and therefore is not selected/counted evidence.
    """
    if policy == POLICY_ALL:
        return True
    if assessment_criterion is None:
        return False
    if policy == POLICY_SELECTED:
        return bool(assessment_criterion.get("selected", True))
    return bool(assessment_criterion.get("counted", True))


def _criterion_lookup(rubric_data: Mapping[str, Any]) -> Tuple[Dict[str, dict], Dict[str, List[dict]]]:
    by_id: Dict[str, dict] = {}
    by_title: Dict[str, List[dict]] = {}
    for criterion in rubric_data.get("criteria", []) or []:
        if not isinstance(criterion, dict):
            continue
        cid = str(criterion.get("id") or "").strip()
        if cid and cid not in by_id:
            by_id[cid] = criterion
        title = str(criterion.get("title") or "").strip()
        if title:
            by_title.setdefault(title, []).append(criterion)
    return by_id, by_title


def _find_rubric_match(
    assessment_criterion: Mapping[str, Any],
    rubric_by_id: Mapping[str, dict],
    rubric_by_title: Mapping[str, Sequence[dict]],
) -> Optional[dict]:
    cid = str(assessment_criterion.get("id") or "").strip()
    if cid:
        return rubric_by_id.get(cid)

    title = str(assessment_criterion.get("title") or "").strip()
    candidates = list(rubric_by_title.get(title, [])) if title else []
    return candidates[0] if len(candidates) == 1 else None


def _find_assessment_match_for_rubric(
    rubric_criterion: Mapping[str, Any],
    assessment_criteria: Sequence[dict],
    used_indexes: set,
) -> Tuple[Optional[dict], Optional[int]]:
    rid = str(rubric_criterion.get("id") or "").strip()
    if rid:
        for index, criterion in enumerate(assessment_criteria):
            if index in used_indexes or not isinstance(criterion, dict):
                continue
            if str(criterion.get("id") or "").strip() == rid:
                return criterion, index

    # Stable IDs are authoritative.  Title fallback is allowed only when one
    # side lacks a stable ID, matching v2.1 partial-save compatibility rules.
    title = str(rubric_criterion.get("title") or "").strip()
    if not title:
        return None, None

    candidates: List[Tuple[int, dict]] = []
    for index, criterion in enumerate(assessment_criteria):
        if index in used_indexes or not isinstance(criterion, dict):
            continue
        if str(criterion.get("title") or "").strip() != title:
            continue
        aid = str(criterion.get("id") or "").strip()
        if not rid or not aid:
            candidates.append((index, criterion))

    if len(candidates) == 1:
        index, criterion = candidates[0]
        return criterion, index
    return None, None


def _metadata_list(
    assessment_criterion: Optional[Mapping[str, Any]],
    rubric_criterion: Optional[Mapping[str, Any]],
    key: str,
    *,
    aliases: Sequence[str] = (),
) -> List[str]:
    """Use persisted assessment metadata first, then rubric fallback."""
    for source in (assessment_criterion, rubric_criterion):
        if not source:
            continue
        for candidate_key in (key, *aliases):
            values = _string_list(source.get(candidate_key))
            if values:
                return values
    return []


def _resolve_question_id(
    assessment_criterion: Optional[Mapping[str, Any]],
    rubric_criterion: Optional[Mapping[str, Any]],
    title: str,
) -> str:
    explicit = _nonblank(
        assessment_criterion.get("question_id") if assessment_criterion else None,
        rubric_criterion.get("question_id") if rubric_criterion else None,
    )
    if explicit:
        return str(explicit).strip()
    # Rubric loading already owns backward-compatible question-ID inference.
    # Keep this exporter reporting-only rather than reimplementing that logic.
    return ""


def _submission_fields(submission_meta: Any) -> Dict[str, str]:
    if not isinstance(submission_meta, Mapping):
        return {
            "submission_source": "",
            "submission_file_latex": "",
            "submission_file_pdf": "",
            "submission_hash_latex": "",
            "submission_hash_pdf": "",
        }

    files = submission_meta.get("files")
    hashes = submission_meta.get("file_hashes")
    if not isinstance(files, Mapping):
        files = {}
    if not isinstance(hashes, Mapping):
        hashes = {}

    # For a normal LaTeX submission, ``compiled_pdf`` is the app-generated
    # human-readable representation.  Prefer an original/reference ``pdf`` if
    # one exists, otherwise expose the compiled PDF in the master evidence row.
    pdf_key = "pdf" if files.get("pdf") else "compiled_pdf"

    return {
        "submission_source": str(submission_meta.get("source_used") or ""),
        "submission_file_latex": str(files.get("latex") or ""),
        "submission_file_pdf": str(files.get(pdf_key) or ""),
        "submission_hash_latex": str(hashes.get("latex_sha256") or ""),
        "submission_hash_pdf": str(
            hashes.get(f"{pdf_key}_sha256")
            or hashes.get("pdf_sha256")
            or ""
        ),
    }


def _performance_band(percentage: Optional[float], outcome_profile: Any) -> str:
    if percentage is None or outcome_profile is None:
        return ""
    bands = getattr(outcome_profile, "performance_bands", None)
    if not isinstance(bands, Mapping):
        return ""
    score = min(max(float(percentage), 0.0), 100.0)
    for name, bounds in bands.items():
        try:
            lo, hi = float(bounds[0]), float(bounds[1])
        except (TypeError, ValueError, IndexError):
            continue
        if lo <= score <= hi:
            return str(name)
    return ""


def _assessment_identity(
    assessment: Mapping[str, Any],
    assessment_file: str,
    warnings: List[dict],
) -> Tuple[str, str]:
    submission_meta = assessment.get("submission_meta")
    if not isinstance(submission_meta, Mapping):
        submission_meta = {}

    student_id = str(_nonblank(
        assessment.get("student_id"),
        submission_meta.get("student_id"),
        default="",
    )).strip()
    student_name = str(assessment.get("student_name") or "").strip()

    if not student_id:
        # Existing roster discovery already uses an assessment filename stem as
        # the compatibility ID for v2.0 files that contain only student_name.
        student_id = Path(assessment_file).stem if assessment_file else ""
        warnings.append(_warning(
            "missing_student_id",
            "Assessment has no student_id; filename stem was used as a compatibility ID.",
            assessment_file=assessment_file,
            student_id=student_id,
        ))
    if not student_name:
        warnings.append(_warning(
            "missing_student_name",
            "Assessment has no student_name; the exported name is blank.",
            assessment_file=assessment_file,
            student_id=student_id,
        ))

    return student_id, student_name


def _build_row(
    *,
    assessment: Mapping[str, Any],
    assessment_criterion: Optional[Mapping[str, Any]],
    rubric_criterion: Optional[Mapping[str, Any]],
    assignment_meta: Mapping[str, Any],
    course_meta: Mapping[str, Any],
    evidence_policy: str,
    outcome_profile: Any,
    assessment_file: str,
    warnings: List[dict],
    student_id: str,
    student_name: str,
) -> dict:
    abet_meta = assessment.get("abet_meta")
    if not isinstance(abet_meta, Mapping):
        abet_meta = {}

    title = str(_nonblank(
        assessment_criterion.get("title") if assessment_criterion else None,
        rubric_criterion.get("title") if rubric_criterion else None,
        default="",
    ))
    criterion_id = str(_nonblank(
        assessment_criterion.get("id") if assessment_criterion else None,
        rubric_criterion.get("id") if rubric_criterion else None,
        default="",
    ))
    description = str(_nonblank(
        assessment_criterion.get("description") if assessment_criterion else None,
        rubric_criterion.get("description") if rubric_criterion else None,
        default="",
    ))

    question_id = _resolve_question_id(assessment_criterion, rubric_criterion, title)
    if not question_id:
        warnings.append(_warning(
            "missing_question_id",
            "Criterion has no explicit or inferable question_id.",
            assessment_file=assessment_file,
            student_id=student_id,
            criterion_id=criterion_id,
        ))

    awarded = _number_or_none(
        assessment_criterion.get("points_awarded") if assessment_criterion else None
    )
    if assessment_criterion is not None and "points_possible" in assessment_criterion:
        possible = _number_or_none(assessment_criterion.get("points_possible"))
    else:
        possible = _number_or_none(
            rubric_criterion.get("points") if rubric_criterion else None
        )

    if awarded is None or possible is None:
        warnings.append(_warning(
            "missing_points",
            "Criterion is missing points_awarded or points_possible.",
            assessment_file=assessment_file,
            student_id=student_id,
            criterion_id=criterion_id,
        ))

    course_outcomes = _metadata_list(
        assessment_criterion, rubric_criterion, "course_outcomes"
    )
    program_outcomes = _metadata_list(
        assessment_criterion,
        rubric_criterion,
        "program_outcomes",
        aliases=("abet_outcomes",),
    )
    abet_outcomes = _metadata_list(
        assessment_criterion,
        rubric_criterion,
        "abet_outcomes",
        aliases=("program_outcomes",),
    )
    assessment_tags = _metadata_list(
        assessment_criterion, rubric_criterion, "assessment_tags"
    )

    if not course_outcomes and not program_outcomes and not abet_outcomes:
        warnings.append(_warning(
            "missing_outcome_mapping",
            "Criterion has no course/program/ABET outcome mapping.",
            assessment_file=assessment_file,
            student_id=student_id,
            criterion_id=criterion_id,
        ))

    selected = _raw_bool(assessment_criterion, "selected")
    counted = _raw_bool(assessment_criterion, "counted")
    if assessment_criterion is not None:
        if "selected" not in assessment_criterion:
            warnings.append(_warning(
                "missing_selected_flag",
                "Criterion has no selected flag; legacy ABET policy semantics apply.",
                assessment_file=assessment_file,
                student_id=student_id,
                criterion_id=criterion_id,
            ))
        if "counted" not in assessment_criterion:
            warnings.append(_warning(
                "missing_counted_flag",
                "Criterion has no counted flag; legacy ABET policy semantics apply.",
                assessment_file=assessment_file,
                student_id=student_id,
                criterion_id=criterion_id,
            ))

    pct = _percentage(awarded, possible)
    submission = _submission_fields(assessment.get("submission_meta"))

    row = {
        "semester": str(_nonblank(
            course_meta.get("semester"),
            assignment_meta.get("semester"),
            abet_meta.get("semester"),
            default="",
        )),
        "course_code": str(_nonblank(
            course_meta.get("course_code"),
            assignment_meta.get("course_code"),
            abet_meta.get("course_code"),
            default="",
        )),
        "course_name": str(_nonblank(
            course_meta.get("course_name"),
            assignment_meta.get("course_name"),
            default="",
        )),
        "section": str(_nonblank(
            course_meta.get("section"),
            assignment_meta.get("section"),
            default="",
        )),
        "assignment_id": str(_nonblank(
            assignment_meta.get("assignment_id"),
            assignment_meta.get("assessment_id"),
            abet_meta.get("assessment_id"),
            default="",
        )),
        "assignment_title": str(_nonblank(
            assignment_meta.get("assignment_title"),
            assignment_meta.get("assessment_name"),
            assessment.get("assignment_name"),
            default="",
        )),
        "assignment_type": str(_nonblank(
            assignment_meta.get("assignment_type"),
            default="",
        )),
        "assignment_date": str(_nonblank(
            assignment_meta.get("assignment_date"),
            default="",
        )),
        "student_id": student_id,
        "student_name": student_name,
        "question_id": question_id,
        "criterion_id": criterion_id,
        "criterion_title": title,
        "criterion_description": description,
        "points_awarded": awarded,
        "points_possible": possible,
        "percentage": pct,
        "selected": selected,
        "counted": counted,
        "evidence_policy": evidence_policy,
        "course_outcomes": course_outcomes,
        "program_outcomes": program_outcomes,
        "abet_outcomes": abet_outcomes,
        "assessment_tags": assessment_tags,
        "performance_band": _performance_band(pct, outcome_profile),
        # Criterion-level target semantics are intentionally not invented.
        "meets_target": None,
        **submission,
        "notes": "",
    }

    # Defensive assertion: this catches accidental schema drift during future
    # commits before an exporter silently emits inconsistent columns.
    if tuple(row.keys()) != MASTER_EVIDENCE_FIELDS:
        raise AssertionError("Master evidence row schema drift detected")
    return row


def build_master_evidence_rows_for_assessment(
    assessment_data: dict,
    rubric_data: dict,
    assignment_meta: Optional[dict] = None,
    course_meta: Optional[dict] = None,
    evidence_policy: str = DEFAULT_POLICY,
    include_excluded: bool = False,
    *,
    outcome_profile: Any = None,
    assessment_file: str = "",
) -> MasterEvidenceBuildResult:
    """Build normalized evidence rows for one saved student assessment.

    This richer per-assessment API is useful for tests and for the assignment
    aggregator.  The design's assignment-level public function remains a
    ``list[dict]`` API.
    """
    if evidence_policy not in VALID_EVIDENCE_POLICIES:
        raise ValueError(
            f"Unsupported evidence policy {evidence_policy!r}; expected one of "
            f"{sorted(VALID_EVIDENCE_POLICIES)}"
        )
    if not isinstance(assessment_data, dict):
        raise ValueError("assessment_data must be a dictionary")
    if not isinstance(rubric_data, dict) or not isinstance(rubric_data.get("criteria"), list):
        raise ValueError("rubric_data must contain a criteria list")

    assignment_meta = assignment_meta or {}
    course_meta = course_meta or {}
    warnings: List[dict] = []
    rows: List[dict] = []

    student_id, student_name = _assessment_identity(
        assessment_data, assessment_file, warnings
    )
    if not isinstance(assessment_data.get("submission_meta"), Mapping):
        warnings.append(_warning(
            "missing_submission_meta",
            "Assessment has no v2.2 submission metadata; submission fields are blank.",
            assessment_file=assessment_file,
            student_id=student_id,
        ))

    assessment_criteria = [
        criterion for criterion in (assessment_data.get("criteria") or [])
        if isinstance(criterion, dict)
    ]
    rubric_criteria = [
        criterion for criterion in (rubric_data.get("criteria") or [])
        if isinstance(criterion, dict)
    ]

    used_assessment_indexes: set = set()

    # Rubric-first iteration guarantees that POLICY_ALL truly means all rubric
    # criteria, even for an old/partial assessment that omitted an untouched
    # criterion.  Any assessment-only criteria are appended afterward.
    for rubric_criterion in rubric_criteria:
        assessment_criterion, match_index = _find_assessment_match_for_rubric(
            rubric_criterion,
            assessment_criteria,
            used_assessment_indexes,
        )
        if match_index is not None:
            used_assessment_indexes.add(match_index)

        include = include_excluded or _criterion_included_for_policy(
            assessment_criterion, evidence_policy
        )
        if not include:
            continue

        row = _build_row(
            assessment=assessment_data,
            assessment_criterion=assessment_criterion,
            rubric_criterion=rubric_criterion,
            assignment_meta=assignment_meta,
            course_meta=course_meta,
            evidence_policy=evidence_policy,
            outcome_profile=outcome_profile,
            assessment_file=assessment_file,
            warnings=warnings,
            student_id=student_id,
            student_name=student_name,
        )
        rows.append(row)

    rubric_by_id, rubric_by_title = _criterion_lookup(rubric_data)

    # Preserve assessment evidence that cannot be found in the supplied rubric.
    # This avoids data loss if a rubric was moved/edited after grading.
    for index, assessment_criterion in enumerate(assessment_criteria):
        if index in used_assessment_indexes:
            continue

        rubric_match = _find_rubric_match(
            assessment_criterion, rubric_by_id, rubric_by_title
        )
        # If there is a valid fallback match not consumed above, use it.  This
        # mainly serves legacy rubrics without IDs.
        if rubric_match is not None:
            rubric_criterion = rubric_match
        else:
            rubric_criterion = None
            warnings.append(_warning(
                "rubric_criterion_not_found",
                "Saved assessment criterion was not found in the supplied rubric.",
                assessment_file=assessment_file,
                student_id=student_id,
                criterion_id=str(assessment_criterion.get("id") or ""),
            ))

        include = include_excluded or _criterion_included_for_policy(
            assessment_criterion, evidence_policy
        )
        if not include:
            continue

        rows.append(_build_row(
            assessment=assessment_data,
            assessment_criterion=assessment_criterion,
            rubric_criterion=rubric_criterion,
            assignment_meta=assignment_meta,
            course_meta=course_meta,
            evidence_policy=evidence_policy,
            outcome_profile=outcome_profile,
            assessment_file=assessment_file,
            warnings=warnings,
            student_id=student_id,
            student_name=student_name,
        ))

    return MasterEvidenceBuildResult(rows=rows, warnings=warnings)


def _assessment_json_paths(assessments_dir: str) -> Iterable[Path]:
    root = Path(assessments_dir).expanduser()
    if not root.exists():
        raise FileNotFoundError(f"Assessments directory not found: {root}")
    if not root.is_dir():
        raise NotADirectoryError(str(root))

    for path in sorted(root.iterdir(), key=lambda p: p.name.casefold()):
        if not path.is_file() or path.suffix.lower() != ".json":
            continue
        # Match existing ABETAssessmentAnalyzer behavior; generated ABET report
        # JSONs are not student assessments.
        if path.name.casefold().startswith("abet"):
            continue
        yield path


def collect_master_evidence_for_assignment(
    rubric_data: dict,
    assessments_dir: str,
    assignment_meta: Optional[dict],
    course_meta: Optional[dict],
    evidence_policy: str = DEFAULT_POLICY,
    include_excluded: bool = False,
    *,
    outcome_profile: Any = None,
) -> MasterEvidenceBuildResult:
    """Build assignment rows and retain warnings for future exporters."""
    if evidence_policy not in VALID_EVIDENCE_POLICIES:
        raise ValueError(
            f"Unsupported evidence policy {evidence_policy!r}; expected one of "
            f"{sorted(VALID_EVIDENCE_POLICIES)}"
        )
    if not isinstance(rubric_data, dict) or not isinstance(rubric_data.get("criteria"), list):
        raise ValueError("rubric_data must contain a criteria list")

    result = MasterEvidenceBuildResult()
    for path in _assessment_json_paths(assessments_dir):
        try:
            with path.open("r", encoding="utf-8") as handle:
                assessment = json.load(handle)
            if not isinstance(assessment, dict) or not isinstance(assessment.get("criteria"), list):
                raise ValueError("root must be an object containing a criteria list")
        except (OSError, json.JSONDecodeError, TypeError, ValueError) as exc:
            result.warnings.append(_warning(
                "assessment_file_unreadable",
                f"Could not read student assessment: {exc}",
                assessment_file=str(path.resolve()),
            ))
            continue

        built = build_master_evidence_rows_for_assessment(
            assessment,
            rubric_data,
            assignment_meta=assignment_meta,
            course_meta=course_meta,
            evidence_policy=evidence_policy,
            include_excluded=include_excluded,
            outcome_profile=outcome_profile,
            assessment_file=str(path.resolve()),
        )
        result.rows.extend(built.rows)
        result.warnings.extend(built.warnings)

    return result


def build_master_evidence_rows_for_assignment(
    rubric_data: dict,
    assessments_dir: str,
    assignment_meta: Optional[dict],
    course_meta: Optional[dict],
    evidence_policy: str = DEFAULT_POLICY,
    include_excluded: bool = False,
    *,
    outcome_profile: Any = None,
) -> List[dict]:
    """Return normalized row dictionaries for one assignment.

    This is the design-doc public API.  Call
    :func:`collect_master_evidence_for_assignment` when warning metadata is also
    needed (semester/export commits will use that richer result).
    """
    return collect_master_evidence_for_assignment(
        rubric_data,
        assessments_dir,
        assignment_meta,
        course_meta,
        evidence_policy=evidence_policy,
        include_excluded=include_excluded,
        outcome_profile=outcome_profile,
    ).rows


# ---------------------------------------------------------------------------
# Semester-level composition (v2.2.1 Commit 2)
# ---------------------------------------------------------------------------


def _semester_entries(semester_config: Mapping[str, Any]) -> List[dict]:
    """Return assignment config entries using the app's current schema first.

    Existing semester configs created by ``tools/create_semester_config.py`` use
    ``assessments``. The v2.2.1 design document uses ``assignments`` in its
    example. Supporting both keeps current configs working without migration.
    """
    raw = semester_config.get("assessments")
    if raw is None:
        raw = semester_config.get("assignments", [])
    if raw is None:
        return []
    if not isinstance(raw, list):
        raise ValueError("semester config assessments/assignments must be a list")
    return [entry for entry in raw if isinstance(entry, dict)]


def _resolve_semester_path(value: Any, base_dir: Optional[str]) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    path = Path(text).expanduser()
    if not path.is_absolute():
        root = Path(base_dir).expanduser() if base_dir else Path.cwd()
        path = root / path
    return str(path.resolve())


def _load_json_object(path: str, *, label: str) -> dict:
    try:
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except FileNotFoundError:
        raise
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Could not read {label}: {exc}") from exc
    if not isinstance(data, dict):
        raise ValueError(f"{label} must contain a JSON object")
    return data


def _semester_course_meta(semester_config: Mapping[str, Any]) -> dict:
    return {
        "semester": str(semester_config.get("semester") or ""),
        "course_code": str(semester_config.get("course_code") or ""),
        "course_name": str(semester_config.get("course_name") or ""),
        "section": str(semester_config.get("section") or ""),
    }


def _assignment_meta_from_entry(
    entry: Mapping[str, Any], rubric_data: Mapping[str, Any]
) -> dict:
    """Normalize current and design-doc assignment metadata names."""
    return {
        "assignment_id": str(_nonblank(
            entry.get("assignment_id"),
            entry.get("assessment_id"),
            rubric_data.get("assessment_id"),
            default="",
        )),
        "assignment_title": str(_nonblank(
            entry.get("assignment_title"),
            entry.get("assessment_title"),
            entry.get("assessment_name"),
            rubric_data.get("title"),
            default="",
        )),
        "assignment_type": str(_nonblank(
            entry.get("assignment_type"),
            entry.get("assessment_type"),
            default="",
        )),
        "assignment_date": str(_nonblank(
            entry.get("assignment_date"),
            entry.get("assessment_date"),
            default="",
        )),
        "weight": entry.get("weight", 1.0),
    }


def _semester_warning(code: str, message: str, assignment_id: str = "") -> dict:
    warning = _warning(code, message)
    warning["assignment_id"] = str(assignment_id or "")
    return warning


def _tag_semester_warning(warning: Mapping[str, Any], assignment_id: str) -> dict:
    tagged = dict(warning)
    if not tagged.get("assignment_id"):
        tagged["assignment_id"] = str(assignment_id or "")
    return tagged


def _load_semester_outcome_profile(
    semester_config: Mapping[str, Any],
    *,
    base_dir: Optional[str],
    warnings: List[dict],
) -> Any:
    profile_id = str(semester_config.get("profile_id") or "").strip()
    if not profile_id:
        return None

    candidate = profile_id
    if profile_id.endswith(".json") or "/" in profile_id or "\\" in profile_id:
        candidate = _resolve_semester_path(profile_id, base_dir)

    try:
        from src.core.outcome_profile import load_profile
        return load_profile(candidate)
    except Exception as exc:
        warnings.append(_warning(
            "outcome_profile_unavailable",
            f"Could not load outcome profile {profile_id!r}: {exc}",
        ))
        return None


def collect_master_evidence_for_semester(
    semester_config: dict,
    evidence_policy: str = DEFAULT_POLICY,
    include_excluded: bool = False,
    *,
    base_dir: Optional[str] = None,
    outcome_profile: Any = None,
) -> MasterEvidenceBuildResult:
    """Compose master evidence rows across configured ABET assignments.

    This reuses Commit 1's assignment-level builder so semester composition does
    not introduce a second interpretation of scoring or selected/counted state.
    Relative paths resolve against ``base_dir``; the config-file helper below
    supplies the config file's parent directory automatically.
    """
    if evidence_policy not in VALID_EVIDENCE_POLICIES:
        raise ValueError(
            f"Unsupported evidence policy {evidence_policy!r}; expected one of "
            f"{sorted(VALID_EVIDENCE_POLICIES)}"
        )
    if not isinstance(semester_config, dict):
        raise ValueError("semester_config must be a dictionary")

    entries = _semester_entries(semester_config)
    result = MasterEvidenceBuildResult()
    if not entries:
        result.warnings.append(_warning(
            "no_assignments_configured",
            "Semester config contains no assessments/assignments.",
        ))
        return result

    if outcome_profile is None:
        outcome_profile = _load_semester_outcome_profile(
            semester_config, base_dir=base_dir, warnings=result.warnings
        )

    semester_course_meta = _semester_course_meta(semester_config)

    for index, entry in enumerate(entries):
        # Match the existing SemesterABETReport boundary.
        if not bool(entry.get("include_in_abet", True)):
            continue

        provisional_id = str(_nonblank(
            entry.get("assignment_id"),
            entry.get("assessment_id"),
            default=f"assignment_{index + 1}",
        ))

        rubric_path = _resolve_semester_path(entry.get("rubric_path"), base_dir)
        if not rubric_path:
            result.warnings.append(_semester_warning(
                "missing_rubric_path",
                "Semester assignment has no rubric_path; assignment was skipped.",
                provisional_id,
            ))
            continue
        if not Path(rubric_path).is_file():
            result.warnings.append(_semester_warning(
                "rubric_file_missing",
                f"Rubric file not found: {rubric_path}",
                provisional_id,
            ))
            continue

        try:
            rubric_data = _load_json_object(rubric_path, label="rubric")
        except ValueError as exc:
            result.warnings.append(_semester_warning(
                "rubric_file_unreadable",
                str(exc),
                provisional_id,
            ))
            continue
        if not isinstance(rubric_data.get("criteria"), list):
            result.warnings.append(_semester_warning(
                "rubric_file_unreadable",
                "Rubric must contain a criteria list.",
                provisional_id,
            ))
            continue

        assignment_meta = _assignment_meta_from_entry(entry, rubric_data)
        assignment_id = assignment_meta.get("assignment_id") or provisional_id

        assessments_value = _nonblank(
            entry.get("assessments_dir"),
            entry.get("assessment_dir"),
            default="",
        )
        assessments_dir = _resolve_semester_path(assessments_value, base_dir)
        if not assessments_dir:
            result.warnings.append(_semester_warning(
                "missing_assessments_dir",
                "Semester assignment has no assessment_dir/assessments_dir; assignment was skipped.",
                assignment_id,
            ))
            continue
        if not Path(assessments_dir).is_dir():
            result.warnings.append(_semester_warning(
                "assessments_dir_missing",
                f"Assessments directory not found: {assessments_dir}",
                assignment_id,
            ))
            continue

        course_meta = dict(semester_course_meta)
        if not course_meta["semester"]:
            course_meta["semester"] = str(rubric_data.get("semester") or "")
        if not course_meta["course_code"]:
            course_meta["course_code"] = str(rubric_data.get("course_code") or "")
        if not course_meta["course_name"]:
            course_meta["course_name"] = str(rubric_data.get("course_name") or "")

        built = collect_master_evidence_for_assignment(
            rubric_data,
            assessments_dir,
            assignment_meta,
            course_meta,
            evidence_policy=evidence_policy,
            include_excluded=include_excluded,
            outcome_profile=outcome_profile,
        )
        result.rows.extend(built.rows)
        result.warnings.extend(
            _tag_semester_warning(warning, assignment_id)
            for warning in built.warnings
        )

    return result


def build_master_evidence_rows_for_semester(
    semester_config: dict,
    evidence_policy: str = DEFAULT_POLICY,
    include_excluded: bool = False,
    *,
    base_dir: Optional[str] = None,
    outcome_profile: Any = None,
) -> List[dict]:
    """Design-doc public API: return normalized rows across a semester config."""
    return collect_master_evidence_for_semester(
        semester_config,
        evidence_policy=evidence_policy,
        include_excluded=include_excluded,
        base_dir=base_dir,
        outcome_profile=outcome_profile,
    ).rows


def collect_master_evidence_for_semester_config(
    config_path: str,
    evidence_policy: str = DEFAULT_POLICY,
    include_excluded: bool = False,
    *,
    outcome_profile: Any = None,
) -> MasterEvidenceBuildResult:
    """Load a semester config JSON and resolve portable relative paths."""
    path = Path(config_path).expanduser().resolve()
    config = _load_json_object(str(path), label="semester config")
    return collect_master_evidence_for_semester(
        config,
        evidence_policy=evidence_policy,
        include_excluded=include_excluded,
        base_dir=str(path.parent),
        outcome_profile=outcome_profile,
    )


def build_master_evidence_rows_for_semester_config(
    config_path: str,
    evidence_policy: str = DEFAULT_POLICY,
    include_excluded: bool = False,
    *,
    outcome_profile: Any = None,
) -> List[dict]:
    """Convenience row-only API for a semester config JSON file."""
    return collect_master_evidence_for_semester_config(
        config_path,
        evidence_policy=evidence_policy,
        include_excluded=include_excluded,
        outcome_profile=outcome_profile,
    ).rows



# ---------------------------------------------------------------------------
# Export layer (v2.2.1 Commit 3)
# ---------------------------------------------------------------------------

# CSV/XLSX use stable human-facing labels while JSON intentionally preserves
# the normalized snake_case row schema used by the backend APIs above.
MASTER_EVIDENCE_COLUMN_LABELS: Tuple[Tuple[str, str], ...] = (
    ("semester", "Semester"),
    ("course_code", "Course Code"),
    ("course_name", "Course Name"),
    ("section", "Section"),
    ("assignment_id", "Assignment ID"),
    ("assignment_title", "Assignment Title"),
    ("assignment_type", "Assignment Type"),
    ("assignment_date", "Assignment Date"),
    ("student_id", "Student ID"),
    ("student_name", "Student Name"),
    ("question_id", "Question ID"),
    ("criterion_id", "Criterion ID"),
    ("criterion_title", "Criterion Title"),
    ("criterion_description", "Criterion Description"),
    ("points_awarded", "Points Awarded"),
    ("points_possible", "Points Possible"),
    ("percentage", "Percentage"),
    ("selected", "Selected"),
    ("counted", "Counted"),
    ("evidence_policy", "Evidence Policy"),
    ("course_outcomes", "Course Outcomes"),
    ("program_outcomes", "Program Outcomes"),
    ("abet_outcomes", "ABET Outcomes"),
    ("assessment_tags", "Assessment Tags"),
    ("performance_band", "Performance Band"),
    ("meets_target", "Meets Target"),
    ("submission_source", "Submission Source"),
    ("submission_file_latex", "Submission Latex File"),
    ("submission_file_pdf", "Submission PDF File"),
    ("submission_hash_latex", "Submission Latex SHA256"),
    ("submission_hash_pdf", "Submission PDF SHA256"),
    ("notes", "Notes"),
)

MASTER_EVIDENCE_EXPORT_COLUMNS: Tuple[str, ...] = tuple(
    label for _, label in MASTER_EVIDENCE_COLUMN_LABELS
)

SUPPORTED_MASTER_EVIDENCE_FORMATS = frozenset({"csv", "xlsx", "json"})

_WARNING_COLUMNS: Tuple[str, ...] = (
    "Code",
    "Message",
    "Assignment ID",
    "Assessment File",
    "Student ID",
    "Criterion ID",
)

_XLSX_HEADER_FILL = "4472C4"
_XLSX_HEADER_FONT = "FFFFFF"


def _export_scalar(value: Any) -> Any:
    """Convert one normalized row value for CSV/XLSX without losing meaning."""
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return ";".join(str(item) for item in value if str(item).strip())
    if isinstance(value, bool):
        return value
    return value


def _tabular_row(row: Mapping[str, Any]) -> dict:
    return {
        label: _export_scalar(row.get(field))
        for field, label in MASTER_EVIDENCE_COLUMN_LABELS
    }


def _student_key(row: Mapping[str, Any]) -> str:
    return str(_nonblank(row.get("student_id"), row.get("student_name"), default=""))


def _row_outcomes(row: Mapping[str, Any]) -> List[str]:
    """Return the unique outcome IDs represented by a criterion row.

    Program and ABET outcomes can be aliases in older/current rubric schemas, so
    a row is counted only once for an outcome ID even if it appears in both.
    """
    outcomes: List[str] = []
    for key in ("course_outcomes", "program_outcomes", "abet_outcomes"):
        for outcome in _string_list(row.get(key)):
            if outcome not in outcomes:
                outcomes.append(outcome)
    return outcomes


def _mean(values: Iterable[Optional[float]]) -> Optional[float]:
    usable = [float(value) for value in values if value is not None]
    if not usable:
        return None
    return sum(usable) / len(usable)


def build_master_evidence_outcome_summary(rows: Sequence[Mapping[str, Any]]) -> List[dict]:
    """Build the XLSX pivot-like summary described by the v2.2.1 design."""
    aggregates: Dict[str, dict] = {}
    for row in rows:
        for outcome in _row_outcomes(row):
            bucket = aggregates.setdefault(outcome, {
                "outcome": outcome,
                "rows": 0,
                "students": set(),
                "total_earned": 0.0,
                "total_possible": 0.0,
                "percentages": [],
            })
            bucket["rows"] += 1
            student = _student_key(row)
            if student:
                bucket["students"].add(student)
            awarded = _number_or_none(row.get("points_awarded"))
            possible = _number_or_none(row.get("points_possible"))
            pct = _number_or_none(row.get("percentage"))
            if awarded is not None:
                bucket["total_earned"] += awarded
            if possible is not None:
                bucket["total_possible"] += possible
            if pct is not None:
                bucket["percentages"].append(pct)

    result: List[dict] = []
    for outcome in sorted(aggregates, key=str.casefold):
        bucket = aggregates[outcome]
        result.append({
            "Outcome": outcome,
            "Rows": bucket["rows"],
            "Students": len(bucket["students"]),
            "Total Earned": bucket["total_earned"],
            "Total Possible": bucket["total_possible"],
            "Average %": _mean(bucket["percentages"]),
        })
    return result


def build_master_evidence_assignment_summary(rows: Sequence[Mapping[str, Any]]) -> List[dict]:
    """Build assignment-level row counts and average criterion percentages."""
    aggregates: Dict[str, dict] = {}
    for row in rows:
        assignment_id = str(row.get("assignment_id") or "")
        assignment_title = str(row.get("assignment_title") or "")
        key = assignment_id or assignment_title or "(unlabeled assignment)"
        bucket = aggregates.setdefault(key, {
            "assignment_id": assignment_id,
            "assignment_title": assignment_title,
            "rows": 0,
            "students": set(),
            "criteria": set(),
            "outcomes": set(),
            "percentages": [],
        })
        bucket["rows"] += 1
        student = _student_key(row)
        if student:
            bucket["students"].add(student)
        criterion = str(row.get("criterion_id") or row.get("criterion_title") or "")
        if criterion:
            bucket["criteria"].add(criterion)
        bucket["outcomes"].update(_row_outcomes(row))
        pct = _number_or_none(row.get("percentage"))
        if pct is not None:
            bucket["percentages"].append(pct)

    result: List[dict] = []
    for key in sorted(aggregates, key=str.casefold):
        bucket = aggregates[key]
        aid = bucket["assignment_id"]
        title = bucket["assignment_title"]
        if aid and title and aid != title:
            display = f"{aid} — {title}"
        else:
            display = aid or title or key
        result.append({
            "Assignment": display,
            "Rows": bucket["rows"],
            "Students": len(bucket["students"]),
            "Criteria": len(bucket["criteria"]),
            "Outcomes Covered": ";".join(sorted(bucket["outcomes"], key=str.casefold)),
            "Average %": _mean(bucket["percentages"]),
        })
    return result


def _master_evidence_summary(rows: Sequence[Mapping[str, Any]]) -> dict:
    students = {_student_key(row) for row in rows if _student_key(row)}
    assignments = {
        str(_nonblank(row.get("assignment_id"), row.get("assignment_title"), default=""))
        for row in rows
    }
    assignments.discard("")
    outcomes = sorted(
        {outcome for row in rows for outcome in _row_outcomes(row)},
        key=str.casefold,
    )
    return {
        "num_rows": len(rows),
        "num_students": len(students),
        "num_assignments": len(assignments),
        "outcomes_covered": outcomes,
    }


def _infer_course_metadata(
    rows: Sequence[Mapping[str, Any]],
    course_meta: Optional[Mapping[str, Any]],
) -> dict:
    supplied = course_meta or {}

    def value_for(key: str) -> str:
        if supplied.get(key) not in (None, ""):
            return str(supplied.get(key))
        for row in rows:
            value = row.get(key)
            if value not in (None, ""):
                return str(value)
        return ""

    return {
        "semester": value_for("semester"),
        "course_code": value_for("course_code"),
        "course_name": value_for("course_name"),
        "section": value_for("section"),
    }


def _infer_evidence_policy(
    rows: Sequence[Mapping[str, Any]], evidence_policy: Optional[str]
) -> str:
    if evidence_policy:
        if evidence_policy not in VALID_EVIDENCE_POLICIES:
            raise ValueError(
                f"Unsupported evidence policy {evidence_policy!r}; expected one of "
                f"{sorted(VALID_EVIDENCE_POLICIES)}"
            )
        return evidence_policy
    policies = {
        str(row.get("evidence_policy") or "").strip()
        for row in rows
        if str(row.get("evidence_policy") or "").strip()
    }
    if len(policies) == 1:
        return next(iter(policies))
    if not policies:
        return DEFAULT_POLICY
    return "mixed"


def _normalize_warning(warning: Mapping[str, Any]) -> dict:
    return {
        "Code": str(warning.get("code") or ""),
        "Message": str(warning.get("message") or ""),
        "Assignment ID": str(warning.get("assignment_id") or ""),
        "Assessment File": str(warning.get("assessment_file") or ""),
        "Student ID": str(warning.get("student_id") or ""),
        "Criterion ID": str(warning.get("criterion_id") or ""),
    }


def _write_master_evidence_csv(rows: Sequence[Mapping[str, Any]], output_dir: Path) -> str:
    path = output_dir / "master_abet_evidence.csv"
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(MASTER_EVIDENCE_EXPORT_COLUMNS))
        writer.writeheader()
        for row in rows:
            writer.writerow(_tabular_row(row))
    return str(path)


def _write_master_evidence_warnings_csv(
    warnings: Sequence[Mapping[str, Any]], output_dir: Path
) -> str:
    path = output_dir / "master_abet_evidence_warnings.csv"
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(_WARNING_COLUMNS))
        writer.writeheader()
        for warning in warnings:
            writer.writerow(_normalize_warning(warning))
    return str(path)


def _write_master_evidence_json(
    rows: Sequence[Mapping[str, Any]],
    warnings: Sequence[Mapping[str, Any]],
    output_dir: Path,
    *,
    course_meta: Optional[Mapping[str, Any]],
    evidence_policy: str,
) -> str:
    path = output_dir / "master_abet_evidence.json"
    payload = {
        "report_type": "master_abet_evidence",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "course": _infer_course_metadata(rows, course_meta),
        "evidence_policy": evidence_policy,
        "rows": [dict(row) for row in rows],
        "warnings": [dict(warning) for warning in warnings],
        "summary": _master_evidence_summary(rows),
    }
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False, default=str)
    return str(path)


def _xlsx_header(ws: Any, columns: Sequence[str]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.append(list(columns))
    for cell in ws[1]:
        cell.font = Font(bold=True, color=_XLSX_HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=_XLSX_HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _xlsx_auto_width(workbook: Any) -> None:
    for ws in workbook.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)


def _write_dict_sheet(ws: Any, columns: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    _xlsx_header(ws, columns)
    for row in rows:
        ws.append([_export_scalar(row.get(column)) for column in columns])
    ws.freeze_panes = "A2"
    if ws.max_column:
        ws.auto_filter.ref = ws.dimensions


def _write_master_evidence_xlsx(
    rows: Sequence[Mapping[str, Any]],
    warnings: Sequence[Mapping[str, Any]],
    output_dir: Path,
    *,
    course_meta: Optional[Mapping[str, Any]],
    evidence_policy: str,
) -> Optional[str]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError:
        return None

    path = output_dir / "master_abet_evidence.xlsx"
    workbook = openpyxl.Workbook()
    evidence_ws = workbook.active
    evidence_ws.title = "Evidence Rows"
    _xlsx_header(evidence_ws, MASTER_EVIDENCE_EXPORT_COLUMNS)
    for row in rows:
        tabular = _tabular_row(row)
        evidence_ws.append([tabular[column] for column in MASTER_EVIDENCE_EXPORT_COLUMNS])
    evidence_ws.freeze_panes = "A2"
    evidence_ws.auto_filter.ref = evidence_ws.dimensions

    # Keep long narrative cells readable without forcing very wide sheets.
    label_to_column = {
        label: index + 1
        for index, label in enumerate(MASTER_EVIDENCE_EXPORT_COLUMNS)
    }
    for label in ("Criterion Description", "Notes"):
        column = label_to_column[label]
        for row_index in range(2, evidence_ws.max_row + 1):
            evidence_ws.cell(row=row_index, column=column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    outcome_ws = workbook.create_sheet("Outcome Summary")
    outcome_rows = build_master_evidence_outcome_summary(rows)
    _write_dict_sheet(
        outcome_ws,
        ["Outcome", "Rows", "Students", "Total Earned", "Total Possible", "Average %"],
        outcome_rows,
    )

    assignment_ws = workbook.create_sheet("Assignment Summary")
    assignment_rows = build_master_evidence_assignment_summary(rows)
    _write_dict_sheet(
        assignment_ws,
        ["Assignment", "Rows", "Students", "Criteria", "Outcomes Covered", "Average %"],
        assignment_rows,
    )

    warnings_ws = workbook.create_sheet("Warnings")
    _xlsx_header(warnings_ws, _WARNING_COLUMNS)
    for warning in warnings:
        normalized = _normalize_warning(warning)
        warnings_ws.append([normalized[column] for column in _WARNING_COLUMNS])
    warnings_ws.freeze_panes = "A2"
    warnings_ws.auto_filter.ref = warnings_ws.dimensions

    readme_ws = workbook.create_sheet("README")
    _xlsx_header(readme_ws, ["Field", "Description"])
    course = _infer_course_metadata(rows, course_meta)
    readme_rows = [
        ("Report", "Master ABET Evidence Export"),
        ("Evidence unit", "One row = student × assignment × rubric criterion."),
        ("Evidence policy", evidence_policy),
        ("Semester", course["semester"]),
        ("Course", " ".join(part for part in [course["course_code"], course["course_name"]] if part)),
        ("Section", course["section"]),
        ("Selected / Counted", "Persisted grading flags. Blank means the historical assessment did not store the flag."),
        ("Outcome/tag lists", "Semicolon-separated in CSV/XLSX; preserved as arrays in JSON."),
        ("Submission metadata", "Optional v2.2 submission provenance. Blank when submission ingestion was not used."),
        ("Performance Band", "Criterion-row band when a profile was available during row generation."),
        ("Meets Target", "Intentionally blank at criterion-row level unless target semantics are explicitly defined."),
        ("Warnings", f"{len(warnings)} non-fatal data-quality warning(s) were recorded."),
    ]
    for field_name, description in readme_rows:
        readme_ws.append([field_name, description])
    for row_index in range(2, readme_ws.max_row + 1):
        readme_ws.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # Percentage columns are numeric values, not preformatted strings.
    percentage_column = label_to_column["Percentage"]
    for row_index in range(2, evidence_ws.max_row + 1):
        cell = evidence_ws.cell(row=row_index, column=percentage_column)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00"
    for ws in (outcome_ws, assignment_ws):
        for cell in ws[1]:
            if cell.value == "Average %":
                col = cell.column
                for row_index in range(2, ws.max_row + 1):
                    target = ws.cell(row=row_index, column=col)
                    if isinstance(target.value, (int, float)):
                        target.number_format = "0.00"

    _xlsx_auto_width(workbook)
    workbook.save(path)
    return str(path)


def export_master_evidence(
    rows: List[dict],
    output_path: str,
    formats: Sequence[str] = ("csv", "xlsx", "json"),
    *,
    warnings: Optional[Sequence[Mapping[str, Any]]] = None,
    course_meta: Optional[Mapping[str, Any]] = None,
    evidence_policy: Optional[str] = None,
) -> Dict[str, Optional[str]]:
    """Export normalized master-evidence rows to CSV, XLSX, and/or JSON.

    ``output_path`` is the destination directory.  The fixed release filenames
    are ``master_abet_evidence.csv``, ``master_abet_evidence.xlsx``, and
    ``master_abet_evidence.json``.  When CSV is requested and warnings exist, a
    companion ``master_abet_evidence_warnings.csv`` is also written.

    XLSX is optional: if ``openpyxl`` is unavailable, the returned ``xlsx``
    value is ``None`` while other requested formats still succeed.
    """
    if not isinstance(rows, list):
        raise ValueError("rows must be a list of normalized evidence dictionaries")
    for index, row in enumerate(rows):
        if not isinstance(row, dict):
            raise ValueError(f"rows[{index}] must be a dictionary")

    requested: List[str] = []
    for raw_format in formats:
        fmt = str(raw_format).strip().lower()
        if not fmt:
            continue
        if fmt not in SUPPORTED_MASTER_EVIDENCE_FORMATS:
            raise ValueError(
                f"Unsupported export format {raw_format!r}; expected one of "
                f"{sorted(SUPPORTED_MASTER_EVIDENCE_FORMATS)}"
            )
        if fmt not in requested:
            requested.append(fmt)
    if not requested:
        raise ValueError("At least one export format is required")

    normalized_warnings = [dict(w) for w in (warnings or [])]
    policy = _infer_evidence_policy(rows, evidence_policy)
    output_dir = Path(output_path).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)
    if not output_dir.is_dir():
        raise NotADirectoryError(str(output_dir))
    output_dir = output_dir.resolve()

    exported: Dict[str, Optional[str]] = {}
    for fmt in requested:
        if fmt == "csv":
            exported["csv"] = _write_master_evidence_csv(rows, output_dir)
            if normalized_warnings:
                exported["warnings_csv"] = _write_master_evidence_warnings_csv(
                    normalized_warnings, output_dir
                )
        elif fmt == "json":
            exported["json"] = _write_master_evidence_json(
                rows,
                normalized_warnings,
                output_dir,
                course_meta=course_meta,
                evidence_policy=policy,
            )
        elif fmt == "xlsx":
            exported["xlsx"] = _write_master_evidence_xlsx(
                rows,
                normalized_warnings,
                output_dir,
                course_meta=course_meta,
                evidence_policy=policy,
            )
    return exported


__all__ = [
    "MASTER_EVIDENCE_FIELDS",
    "MASTER_EVIDENCE_COLUMN_LABELS",
    "MASTER_EVIDENCE_EXPORT_COLUMNS",
    "SUPPORTED_MASTER_EVIDENCE_FORMATS",
    "MasterEvidenceBuildResult",
    "VALID_EVIDENCE_POLICIES",
    "build_master_evidence_rows_for_assessment",
    "build_master_evidence_rows_for_assignment",
    "collect_master_evidence_for_assignment",
    "build_master_evidence_rows_for_semester",
    "collect_master_evidence_for_semester",
    "build_master_evidence_rows_for_semester_config",
    "collect_master_evidence_for_semester_config",
    "build_master_evidence_outcome_summary",
    "build_master_evidence_assignment_summary",
    "export_master_evidence",
]
