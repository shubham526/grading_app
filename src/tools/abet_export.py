"""
ABET Export Engine — Rubric Grading Tool
=========================================

Exports assignment-level ABET reports to:
  - abet_report.json
  - abet_assignment_summary.csv
  - abet_student_outcomes.csv
  - abet_criterion_breakdown.csv
  - abet_unmapped_criteria.csv
  - abet_assignment_summary.xlsx  (multi-sheet workbook)
"""

from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from typing import Dict, List, Optional


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_assignment_report(
    report: dict,
    output_dir: str,
    include_xlsx: bool = True,
) -> List[str]:
    """
    Write all assignment-report output files to output_dir.

    Args:
        report:       The report dict produced by ABETAssessmentAnalyzer.
        output_dir:   Directory to write files into (created if needed).
        include_xlsx: Whether to write the Excel workbook.

    Returns:
        List of file paths written.
    """
    os.makedirs(output_dir, exist_ok=True)
    written: List[str] = []

    written.append(_write_json(report, output_dir))
    written.extend(_write_assignment_csvs(report, output_dir))
    if include_xlsx:
        path = _write_assignment_xlsx(report, output_dir)
        if path:
            written.append(path)

    return written


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

def _write_json(report: dict, output_dir: str) -> str:
    path = os.path.join(output_dir, "abet_report.json")
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2, default=str)
    return path


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

def _write_assignment_csvs(report: dict, output_dir: str) -> List[str]:
    written: List[str] = []
    written.append(_write_outcome_summary_csv(report, output_dir, "course_outcomes", "lo"))
    # Accept both canonical name and alias
    po_key = "program_outcomes" if "program_outcomes" in report.get("outcome_summary", {}) \
             else "abet_outcomes"
    written.append(_write_outcome_summary_csv(report, output_dir, po_key, "so"))
    written.append(_write_student_outcomes_csv(report, output_dir))
    written.append(_write_criterion_breakdown_csv(report, output_dir))
    written.append(_write_unmapped_criteria_csv(report, output_dir))
    return [p for p in written if p]


def _write_outcome_summary_csv(report: dict, output_dir: str,
                                key: str, suffix: str) -> Optional[str]:
    data = report.get("outcome_summary", {}).get(key, {})
    if not data:
        return None

    path = os.path.join(output_dir, f"abet_assignment_{suffix}_summary.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow([suffix.upper(), "Description", "Evidence Items",
                         "Mean %", "Median %", "Std Dev", "N Students",
                         "Adequate+ %", "Meets Target"])

        course_info = report.get("course_info", {})
        target = course_info.get("target_percentage", 70.0)
        descriptions = report.get("outcome_descriptions", {}).get(key, {})

        for oid in sorted(data.keys()):
            d = data[oid]
            # Fix 1: band_counts and proficient_plus_pct now live in outcome_summary
            # (back-filled by check_targets), so no separate lookup needed.
            adeq_pct  = d.get("proficient_plus_pct", 0.0)
            meets     = d.get("meets_target", adeq_pct >= target)
            writer.writerow([
                oid,
                descriptions.get(oid, ""),
                d.get("count", 0),
                f"{d.get('mean', 0):.1f}",
                f"{d.get('median', 0):.1f}",
                f"{d.get('std_dev', 0):.1f}",
                d.get("count", 0),
                f"{adeq_pct:.1f}",
                "Yes" if meets else "No",
            ])
    return path


def _write_student_outcomes_csv(report: dict, output_dir: str) -> Optional[str]:
    rows = report.get("student_outcome_scores", [])
    if not rows:
        return None

    path = os.path.join(output_dir, "abet_student_outcomes.csv")
    if not rows:
        return None

    # Collect all outcome IDs
    all_lo = sorted({k for r in rows for k in r.get("lo_scores", {})})
    all_so = sorted({k for r in rows for k in r.get("so_scores", {})})

    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Student"] + all_lo + all_so)
        for row in rows:
            name = row.get("student_name", "")
            lo_vals = [f"{row.get('lo_scores', {}).get(lo, ''):.1f}"
                       if row.get('lo_scores', {}).get(lo) is not None else ""
                       for lo in all_lo]
            so_vals = [f"{row.get('so_scores', {}).get(so, ''):.1f}"
                       if row.get('so_scores', {}).get(so) is not None else ""
                       for so in all_so]
            writer.writerow([name] + lo_vals + so_vals)
    return path


def _write_criterion_breakdown_csv(report: dict, output_dir: str) -> Optional[str]:
    rows = report.get("criterion_outcome_breakdown", [])
    if not rows:
        return None
    path = os.path.join(output_dir, "abet_criterion_breakdown.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Criterion ID", "Title", "Points",
                         "Course LOs", "ABET SOs", "Students Counted", "Mean %"])
        for row in rows:
            writer.writerow([
                row.get("id", ""),
                row.get("title", ""),
                row.get("points", ""),
                ", ".join(row.get("course_outcomes", [])),
                ", ".join(row.get("abet_outcomes", [])),
                row.get("students_counted", ""),
                f"{row.get('mean_pct', 0):.1f}" if row.get("mean_pct") is not None else "",
            ])
    return path


def _write_unmapped_criteria_csv(report: dict, output_dir: str) -> Optional[str]:
    rows = report.get("unmapped_criteria", [])
    if not rows:
        return None
    path = os.path.join(output_dir, "abet_unmapped_criteria.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Criterion ID", "Title", "Points", "Selected?", "Counted?"])
        for row in rows:
            writer.writerow([
                row.get("id", ""),
                row.get("title", ""),
                row.get("points_possible", ""),
                "Yes" if row.get("selected") else "No",
                "Yes" if row.get("counted") else "No",
            ])
    return path


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------

def _write_assignment_xlsx(report: dict, output_dir: str) -> Optional[str]:
    """Write a 6-sheet Excel workbook for the assignment-level report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[abet_export] openpyxl not installed — skipping XLSX export")
        return None

    path = os.path.join(output_dir, "abet_assignment_summary.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    course_info = report.get("course_info", {})
    target = course_info.get("target_percentage", 70.0)
    outcome_desc = report.get("outcome_descriptions", {})
    outcome_summary = report.get("outcome_summary", {})

    # --- Sheet 1: Course LO Summary ---
    ws = wb.create_sheet("Course LO Summary")
    _xlsx_header(ws, ["LO", "Description", "Evidence Items",
                       "Mean %", "Median %", "Std Dev", "N", "Adequate+ %", "Meets Target"])
    lo_data = outcome_summary.get("course_outcomes", {})
    lo_desc = outcome_desc.get("course_outcomes", {})
    for oid in sorted(lo_data.keys()):
        d = lo_data[oid]
        adeq_pct = d.get("proficient_plus_pct", 0.0)
        meets    = d.get("meets_target", adeq_pct >= target)
        ws.append([oid, lo_desc.get(oid, ""), d.get("count", 0),
                   round(d.get("mean", 0), 1), round(d.get("median", 0), 1),
                   round(d.get("std_dev", 0), 1), d.get("count", 0),
                   round(adeq_pct, 1), "Yes" if meets else "No"])

    # --- Sheet 2: Program/ABET SO Summary ---
    ws2 = wb.create_sheet("ABET SO Summary")
    _xlsx_header(ws2, ["SO", "Description", "Evidence Items",
                        "Mean %", "Median %", "Std Dev", "N", "Adequate+ %", "Meets Target"])
    # Accept both canonical and alias key
    so_data = (outcome_summary.get("program_outcomes") or
               outcome_summary.get("abet_outcomes") or {})
    so_desc = (outcome_desc.get("program_outcomes") or
               outcome_desc.get("abet_outcomes") or {})
    for oid in sorted(so_data.keys()):
        d = so_data[oid]
        adeq_pct = d.get("proficient_plus_pct", 0.0)
        meets    = d.get("meets_target", adeq_pct >= target)
        ws2.append([oid, so_desc.get(oid, ""), d.get("count", 0),
                    round(d.get("mean", 0), 1), round(d.get("median", 0), 1),
                    round(d.get("std_dev", 0), 1), d.get("count", 0),
                    round(adeq_pct, 1), "Yes" if meets else "No"])

    # --- Sheet 3: Criterion Breakdown ---
    ws3 = wb.create_sheet("Criterion Breakdown")
    _xlsx_header(ws3, ["Criterion ID", "Title", "Points",
                        "Course LOs", "ABET SOs", "Students Counted", "Mean %"])
    for row in report.get("criterion_outcome_breakdown", []):
        ws3.append([row.get("id", ""), row.get("title", ""), row.get("points", ""),
                    ", ".join(row.get("course_outcomes", [])),
                    ", ".join(row.get("abet_outcomes", [])),
                    row.get("students_counted", ""),
                    round(row.get("mean_pct", 0) or 0, 1)])

    # --- Sheet 4: Student Outcome Scores ---
    ws4 = wb.create_sheet("Student Outcome Scores")
    student_rows = report.get("student_outcome_scores", [])
    all_lo = sorted({k for r in student_rows for k in r.get("lo_scores", {})})
    all_so = sorted({k for r in student_rows for k in r.get("so_scores", {})})
    _xlsx_header(ws4, ["Student"] + all_lo + all_so)
    for row in student_rows:
        lo_scores = row.get("lo_scores", {})
        so_scores = row.get("so_scores", {})
        # Use None for missing outcomes so the cell is blank, not zero (reviewer fix)
        lo_vals = [round(lo_scores[lo], 1) if lo in lo_scores else None for lo in all_lo]
        so_vals = [round(so_scores[so], 1) if so in so_scores else None for so in all_so]
        ws4.append([row.get("student_name", "")] + lo_vals + so_vals)

    # --- Sheet 5: Unmapped Criteria ---
    ws5 = wb.create_sheet("Unmapped Criteria")
    _xlsx_header(ws5, ["Criterion ID", "Title", "Points", "Selected?", "Counted?"])
    for row in report.get("unmapped_criteria", []):
        ws5.append([row.get("id", ""), row.get("title", ""),
                    row.get("points_possible", ""),
                    "Yes" if row.get("selected") else "No",
                    "Yes" if row.get("counted") else "No"])

    # --- Sheet 6: Mapping Warnings ---
    ws6 = wb.create_sheet("Mapping Warnings")
    _xlsx_header(ws6, ["Level", "Code", "Message", "Criterion ID"])
    for issue in report.get("mapping_warnings", []):
        ws6.append([issue.get("level", ""), issue.get("code", ""),
                    issue.get("message", ""), issue.get("criterion_id", "")])

    _xlsx_auto_width(wb)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# openpyxl helpers
# ---------------------------------------------------------------------------

_HEADER_FILL  = "4472C4"
_HEADER_FONT  = "FFFFFF"


def _xlsx_header(ws, columns: List[str]) -> None:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        ws.append(columns)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color=_HEADER_FONT)
            cell.fill      = PatternFill("solid", fgColor=_HEADER_FILL)
            cell.alignment = Alignment(horizontal="center")
    except Exception:
        ws.append(columns)


def _xlsx_auto_width(wb) -> None:
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


# ===========================================================================
# SEMESTER REPORT EXPORT
# ===========================================================================

def export_semester_report(
    report: dict,
    output_dir: str,
    include_xlsx: bool = True,
) -> List[str]:
    """
    Write all semester-report output files to output_dir.

    Files written:
      abet_report.json
      abet_semester_summary.csv
      abet_outcome_by_assessment.csv
      abet_evidence_coverage.csv
      abet_semester_summary.xlsx   (7-sheet workbook, if openpyxl available)

    Returns list of file paths written.
    """
    os.makedirs(output_dir, exist_ok=True)
    written: List[str] = []

    written.append(_write_json(report, output_dir))
    written.extend(_write_semester_csvs(report, output_dir))
    if include_xlsx:
        path = _write_semester_xlsx(report, output_dir)
        if path:
            written.append(path)

    return [p for p in written if p]


# ---------------------------------------------------------------------------
# Semester CSV helpers
# ---------------------------------------------------------------------------

def _write_semester_csvs(report: dict, output_dir: str) -> List[str]:
    written: List[str] = []
    written.append(_write_semester_summary_csv(report, output_dir))
    written.append(_write_outcome_by_assessment_csv(report, output_dir))
    written.append(_write_evidence_coverage_csv(report, output_dir))
    written.append(_write_semester_student_outcomes_csv(report, output_dir))
    written.append(_write_semester_unmapped_csv(report, output_dir))
    return [p for p in written if p]


def _write_semester_summary_csv(report: dict, output_dir: str) -> Optional[str]:
    summary  = report.get("semester_summary", {})
    po_data  = summary.get("program_outcomes", {})
    if not po_data:
        return None

    target   = float(report.get("course_info", {}).get("target_percentage", 75.0))
    po_desc  = report.get("outcome_descriptions", {}).get("program_outcomes", {})
    path     = os.path.join(output_dir, "abet_semester_summary.csv")

    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Outcome", "Description", "Evidence Sources",
                         "Mean %", "Median %", "Std Dev", "N",
                         "Adequate+ %", "Meets Target"])
        for oid in sorted(po_data.keys()):
            d        = po_data[oid]
            adeq_pct = d.get("proficient_plus_pct", 0.0)
            meets    = d.get("meets_target", adeq_pct >= target)
            # Count distinct assessments that provided evidence for this outcome
            sources = sum(
                1 for asmnt in report.get("assignment_details", [])
                if oid in asmnt.get("so_covered", [])
            )
            writer.writerow([
                oid, po_desc.get(oid, ""), sources,
                f"{d.get('mean', 0):.1f}", f"{d.get('median', 0):.1f}",
                f"{d.get('std_dev', 0):.1f}", d.get("count", 0),
                f"{adeq_pct:.1f}", "Yes" if meets else "No",
            ])
    return path


def _write_outcome_by_assessment_csv(report: dict, output_dir: str) -> Optional[str]:
    summary     = report.get("semester_summary", {})
    by_asmnt_so = summary.get("by_assessment_so", {})
    details     = report.get("assignment_details", [])
    if not by_asmnt_so or not details:
        return None

    asmnt_names = [a.get("assessment_name", a.get("assessment_id", ""))
                   for a in details if "error" not in a]
    all_sos     = sorted(by_asmnt_so.keys())

    path = os.path.join(output_dir, "abet_outcome_by_assessment.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Outcome"] + asmnt_names + ["Overall Mean"])
        for oid in all_sos:
            row = [oid]
            per_asmnt = by_asmnt_so.get(oid, {})
            for name in asmnt_names:
                val = per_asmnt.get(name)
                row.append(f"{val:.1f}" if val is not None else "—")
            overall = summary.get("program_outcomes", {}).get(oid, {}).get("mean")
            row.append(f"{overall:.1f}" if overall is not None else "—")
            writer.writerow(row)
    return path


def _write_evidence_coverage_csv(report: dict, output_dir: str) -> Optional[str]:
    summary  = report.get("semester_summary", {})
    coverage = summary.get("coverage_matrix", {})
    all_los  = summary.get("all_lo_ids", [])
    if not coverage or not all_los:
        return None

    path = os.path.join(output_dir, "abet_evidence_coverage.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Assessment"] + all_los)
        for asmnt_name, lo_map in sorted(coverage.items()):
            row = [asmnt_name] + ["Yes" if lo_map.get(lo) else "No"
                                   for lo in all_los]
            writer.writerow(row)
    return path


def _write_semester_student_outcomes_csv(report: dict, output_dir: str) -> Optional[str]:
    """Per-student outcomes aggregated across all assignments."""
    details = report.get("assignment_details", [])
    all_rows: List[dict] = []
    for asmnt in details:
        name = asmnt.get("assessment_name", "")
        # student_outcome_scores may not be present at semester level
        for student in asmnt.get("student_outcome_scores", []):
            row = dict(student)
            row["assessment"] = name
            all_rows.append(row)
    if not all_rows:
        return None

    all_lo = sorted({k for r in all_rows for k in r.get("lo_scores", {})})
    all_so = sorted({k for r in all_rows for k in r.get("so_scores", {})})
    path   = os.path.join(output_dir, "abet_student_outcomes_all.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Assessment", "Student"] + all_lo + all_so)
        for row in all_rows:
            lo_vals = [f"{row.get('lo_scores',{}).get(lo,''):.1f}"
                       if row.get('lo_scores',{}).get(lo) is not None else ""
                       for lo in all_lo]
            so_vals = [f"{row.get('so_scores',{}).get(so,''):.1f}"
                       if row.get('so_scores',{}).get(so) is not None else ""
                       for so in all_so]
            writer.writerow([row.get("assessment",""), row.get("student_name","")]
                            + lo_vals + so_vals)
    return path


def _write_semester_unmapped_csv(report: dict, output_dir: str) -> Optional[str]:
    rows: List[dict] = []
    for asmnt in report.get("assignment_details", []):
        for u in asmnt.get("unmapped_criteria", []):
            u2 = dict(u)
            u2["assessment"] = asmnt.get("assessment_name", "")
            rows.append(u2)
    if not rows:
        return None

    path = os.path.join(output_dir, "abet_unmapped_criteria.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Assessment", "Criterion ID", "Title", "Points"])
        for row in rows:
            writer.writerow([row.get("assessment",""), row.get("id",""),
                             row.get("title",""), row.get("points_possible","")])
    return path


# ---------------------------------------------------------------------------
# Semester XLSX
# ---------------------------------------------------------------------------

def _write_semester_xlsx(report: dict, output_dir: str) -> Optional[str]:
    try:
        import openpyxl
    except ImportError:
        print("[abet_export] openpyxl not installed — skipping semester XLSX")
        return None

    path     = os.path.join(output_dir, "abet_semester_summary.xlsx")
    wb       = openpyxl.Workbook()
    wb.remove(wb.active)

    summary     = report.get("semester_summary", {})
    details     = report.get("assignment_details", [])
    po_data     = summary.get("program_outcomes", {})
    lo_data     = summary.get("course_outcomes", {})
    po_desc     = report.get("outcome_descriptions", {}).get("program_outcomes", {})
    lo_desc     = report.get("outcome_descriptions", {}).get("course_outcomes", {})
    course_info = report.get("course_info", {})
    target      = float(course_info.get("target_percentage", 75.0))
    asmnt_names = [a.get("assessment_name", a.get("assessment_id", ""))
                   for a in details if "error" not in a]

    # --- Sheet 1: Semester Summary (Program Outcomes) ---
    ws1 = wb.create_sheet("Semester Summary")
    _xlsx_header(ws1, ["Outcome", "Description", "Evidence Sources",
                        "Mean %", "Median %", "Std Dev", "N",
                        "Adequate+ %", "Meets Target"])
    for oid in sorted(po_data.keys()):
        d        = po_data[oid]
        adeq_pct = d.get("proficient_plus_pct", 0.0)
        meets    = d.get("meets_target", adeq_pct >= target)
        sources  = sum(1 for a in details if oid in a.get("so_covered", []))
        ws1.append([oid, po_desc.get(oid, ""), sources,
                    round(d.get("mean",0),1), round(d.get("median",0),1),
                    round(d.get("std_dev",0),1), d.get("count",0),
                    round(adeq_pct,1), "Yes" if meets else "No"])

    # --- Sheet 2: Outcome by Assessment ---
    ws2 = wb.create_sheet("Outcome by Assessment")
    by_so = summary.get("by_assessment_so", {})
    _xlsx_header(ws2, ["Outcome"] + asmnt_names + ["Overall Mean"])
    for oid in sorted(by_so.keys()):
        per = by_so.get(oid, {})
        row = [oid] + [round(per.get(n),1) if per.get(n) is not None else None
                       for n in asmnt_names]
        overall = po_data.get(oid, {}).get("mean")
        row.append(round(overall,1) if overall is not None else None)
        ws2.append(row)

    # --- Sheet 3: Evidence Coverage ---
    ws3 = wb.create_sheet("Evidence Coverage")
    all_los  = summary.get("all_lo_ids", [])
    coverage = summary.get("coverage_matrix", {})
    _xlsx_header(ws3, ["Assessment"] + all_los)
    for aname in sorted(coverage.keys()):
        lo_map = coverage[aname]
        ws3.append([aname] + ["Yes" if lo_map.get(lo) else "No"
                               for lo in all_los])

    # --- Sheet 4: Assessment Details ---
    ws4 = wb.create_sheet("Assessment Details")
    _xlsx_header(ws4, ["Assessment", "Assessment ID", "Students",
                        "Weight", "Program Outcomes", "Course LOs"])
    for a in details:
        ws4.append([
            a.get("assessment_name",""),
            a.get("assessment_id",""),
            a.get("num_students",0),
            a.get("weight",1.0),
            ", ".join(a.get("so_covered",[])),
            ", ".join(a.get("lo_covered",[])),
        ])

    # --- Sheet 5: Course LO Summary ---
    ws5 = wb.create_sheet("Course LO Summary")
    _xlsx_header(ws5, ["LO", "Description", "Mean %", "Median %",
                        "Std Dev", "N", "Adequate+ %", "Meets Target"])
    for oid in sorted(lo_data.keys()):
        d        = lo_data[oid]
        adeq_pct = d.get("proficient_plus_pct", 0.0)
        meets    = d.get("meets_target", adeq_pct >= target)
        ws5.append([oid, lo_desc.get(oid,""),
                    round(d.get("mean",0),1), round(d.get("median",0),1),
                    round(d.get("std_dev",0),1), d.get("count",0),
                    round(adeq_pct,1), "Yes" if meets else "No"])

    # --- Sheet 6: Unmapped Criteria ---
    ws6 = wb.create_sheet("Unmapped Criteria")
    _xlsx_header(ws6, ["Assessment", "Criterion ID", "Title", "Points"])
    for a in details:
        for u in a.get("unmapped_criteria", []):
            ws6.append([a.get("assessment_name",""),
                        u.get("id",""), u.get("title",""),
                        u.get("points_possible","")])

    # --- Sheet 7: Closing-the-Loop Notes ---
    ws7 = wb.create_sheet("Notes")
    _xlsx_header(ws7, ["Field", "Content"])
    ctl = report.get("closing_the_loop", {})
    for field, label in [
        ("reflection",              "Reflection"),
        ("planned_improvements",    "Planned Improvements"),
        ("notes_for_next_offering", "Notes for Next Offering"),
    ]:
        ws7.append([label, ctl.get(field, "")])

    _xlsx_auto_width(wb)
    wb.save(path)
    return path
