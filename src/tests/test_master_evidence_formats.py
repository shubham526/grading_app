"""Tests for v2.2.1 Commit 3 master ABET evidence export formats."""

import csv
import json
import tempfile
import unittest
from pathlib import Path

from src.tools.master_evidence_export import (
    MASTER_EVIDENCE_EXPORT_COLUMNS,
    MASTER_EVIDENCE_FIELDS,
    build_master_evidence_assignment_summary,
    build_master_evidence_outcome_summary,
    export_master_evidence,
)

try:
    import openpyxl
except ImportError:  # pragma: no cover - exercised on installations without XLSX support
    openpyxl = None


def _row(**overrides):
    row = {
        "semester": "Fall 2026",
        "course_code": "CS 2500",
        "course_name": "Algorithms",
        "section": "104",
        "assignment_id": "PS1",
        "assignment_title": "Asymptotic Analysis",
        "assignment_type": "problem_set",
        "assignment_date": "2026-09-05",
        "student_id": "alice",
        "student_name": "Alice Smith",
        "question_id": "Q1",
        "criterion_id": "PS1_Q1_RUNTIME",
        "criterion_title": "Question 1 - Runtime",
        "criterion_description": "Analyze runtime, including the recurrence.\nShow work.",
        "points_awarded": 8.0,
        "points_possible": 10.0,
        "percentage": 80.0,
        "selected": True,
        "counted": True,
        "evidence_policy": "counted_only",
        "course_outcomes": ["LO1", "LO4"],
        "program_outcomes": ["SO1", "SO6"],
        "abet_outcomes": ["SO1", "SO6"],
        "assessment_tags": ["runtime", "proof"],
        "performance_band": "adequate",
        "meets_target": None,
        "submission_source": "latex",
        "submission_file_latex": "/evidence/alice/main.tex",
        "submission_file_pdf": "/evidence/alice/main.pdf",
        "submission_hash_latex": "latex-hash",
        "submission_hash_pdf": "pdf-hash",
        "notes": "Instructor note, with comma",
    }
    row.update(overrides)
    assert tuple(row.keys()) == MASTER_EVIDENCE_FIELDS
    return row


def _warnings():
    return [{
        "code": "missing_submission_meta",
        "message": "Assessment has no submission metadata.",
        "assignment_id": "PS2",
        "assessment_file": "/tmp/bob.json",
        "student_id": "bob",
        "criterion_id": "PS2_Q1",
    }]


class TestMasterEvidenceCSV(unittest.TestCase):

    def test_csv_writes_stable_required_column_order(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = export_master_evidence([_row()], tmp, formats=["csv"])
            with open(paths["csv"], newline="", encoding="utf-8") as handle:
                reader = csv.reader(handle)
                header = next(reader)
            self.assertEqual(tuple(header), MASTER_EVIDENCE_EXPORT_COLUMNS)

    def test_csv_semicolon_serializes_outcomes_and_tags(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence([_row()], tmp, formats=["csv"])["csv"]
            with open(path, newline="", encoding="utf-8") as handle:
                row = next(csv.DictReader(handle))
            self.assertEqual(row["Course Outcomes"], "LO1;LO4")
            self.assertEqual(row["Program Outcomes"], "SO1;SO6")
            self.assertEqual(row["ABET Outcomes"], "SO1;SO6")
            self.assertEqual(row["Assessment Tags"], "runtime;proof")

    def test_csv_quotes_commas_and_newlines_safely(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence([_row()], tmp, formats=["csv"])["csv"]
            with open(path, newline="", encoding="utf-8") as handle:
                row = next(csv.DictReader(handle))
            self.assertEqual(
                row["Criterion Description"],
                "Analyze runtime, including the recurrence.\nShow work.",
            )
            self.assertEqual(row["Notes"], "Instructor note, with comma")

    def test_csv_preserves_missing_booleans_as_blank(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence(
                [_row(selected=None, counted=None)], tmp, formats=["csv"]
            )["csv"]
            with open(path, newline="", encoding="utf-8") as handle:
                row = next(csv.DictReader(handle))
            self.assertEqual(row["Selected"], "")
            self.assertEqual(row["Counted"], "")

    def test_warning_csv_is_written_when_warnings_are_supplied(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = export_master_evidence(
                [_row()], tmp, formats=["csv"], warnings=_warnings()
            )
            self.assertIn("warnings_csv", paths)
            with open(paths["warnings_csv"], newline="", encoding="utf-8") as handle:
                warning = next(csv.DictReader(handle))
            self.assertEqual(warning["Code"], "missing_submission_meta")
            self.assertEqual(warning["Assignment ID"], "PS2")
            self.assertEqual(warning["Student ID"], "bob")

    def test_warning_csv_is_not_created_without_warnings(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = export_master_evidence([_row()], tmp, formats=["csv"])
            self.assertNotIn("warnings_csv", paths)
            self.assertFalse(Path(tmp, "master_abet_evidence_warnings.csv").exists())


class TestMasterEvidenceJSON(unittest.TestCase):

    def test_json_has_design_doc_envelope_and_preserves_lists(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence(
                [_row()], tmp, formats=["json"], warnings=_warnings()
            )["json"]
            with open(path, encoding="utf-8") as handle:
                payload = json.load(handle)
            self.assertEqual(payload["report_type"], "master_abet_evidence")
            self.assertTrue(payload["generated_at"])
            self.assertEqual(payload["course"]["course_code"], "CS 2500")
            self.assertEqual(payload["evidence_policy"], "counted_only")
            self.assertEqual(payload["rows"][0]["course_outcomes"], ["LO1", "LO4"])
            self.assertEqual(payload["warnings"][0]["code"], "missing_submission_meta")

    def test_json_summary_counts_rows_students_assignments_and_outcomes(self):
        rows = [
            _row(),
            _row(
                assignment_id="PS2",
                assignment_title="Proofs",
                student_id="bob",
                student_name="Bob Jones",
                criterion_id="PS2_Q1_PROOF",
                course_outcomes=["LO2"],
                program_outcomes=["SO1"],
                abet_outcomes=["SO1"],
            ),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence(rows, tmp, formats=["json"])["json"]
            payload = json.loads(Path(path).read_text(encoding="utf-8"))
            self.assertEqual(payload["summary"]["num_rows"], 2)
            self.assertEqual(payload["summary"]["num_students"], 2)
            self.assertEqual(payload["summary"]["num_assignments"], 2)
            self.assertEqual(
                payload["summary"]["outcomes_covered"],
                ["LO1", "LO2", "LO4", "SO1", "SO6"],
            )

    def test_empty_rows_can_export_course_metadata_supplied_explicitly(self):
        course = {
            "semester": "Fall 2026",
            "course_code": "CS 2500",
            "course_name": "Algorithms",
            "section": "104",
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence(
                [],
                tmp,
                formats=["json"],
                course_meta=course,
                evidence_policy="all",
            )["json"]
            payload = json.loads(Path(path).read_text(encoding="utf-8"))
            self.assertEqual(payload["course"], course)
            self.assertEqual(payload["evidence_policy"], "all")
            self.assertEqual(payload["summary"]["num_rows"], 0)


class TestMasterEvidenceSummaries(unittest.TestCase):

    def test_outcome_summary_deduplicates_program_and_abet_aliases_per_row(self):
        summary = build_master_evidence_outcome_summary([_row()])
        by_outcome = {entry["Outcome"]: entry for entry in summary}
        self.assertEqual(by_outcome["SO1"]["Rows"], 1)
        self.assertEqual(by_outcome["SO6"]["Rows"], 1)
        self.assertEqual(by_outcome["LO1"]["Rows"], 1)

    def test_outcome_summary_aggregates_students_points_and_average(self):
        rows = [
            _row(student_id="alice", points_awarded=8, points_possible=10, percentage=80),
            _row(student_id="bob", points_awarded=6, points_possible=10, percentage=60),
        ]
        summary = build_master_evidence_outcome_summary(rows)
        lo1 = next(item for item in summary if item["Outcome"] == "LO1")
        self.assertEqual(lo1["Rows"], 2)
        self.assertEqual(lo1["Students"], 2)
        self.assertEqual(lo1["Total Earned"], 14.0)
        self.assertEqual(lo1["Total Possible"], 20.0)
        self.assertEqual(lo1["Average %"], 70.0)

    def test_assignment_summary_aggregates_rows_students_criteria_outcomes(self):
        rows = [
            _row(),
            _row(
                student_id="bob",
                criterion_id="PS1_Q2_PROOF",
                question_id="Q2",
                percentage=60,
                course_outcomes=["LO2"],
                program_outcomes=["SO1"],
                abet_outcomes=["SO1"],
            ),
        ]
        summary = build_master_evidence_assignment_summary(rows)
        self.assertEqual(len(summary), 1)
        item = summary[0]
        self.assertEqual(item["Assignment"], "PS1 — Asymptotic Analysis")
        self.assertEqual(item["Rows"], 2)
        self.assertEqual(item["Students"], 2)
        self.assertEqual(item["Criteria"], 2)
        self.assertEqual(item["Outcomes Covered"], "LO1;LO2;LO4;SO1;SO6")
        self.assertEqual(item["Average %"], 70.0)


@unittest.skipUnless(openpyxl is not None, "openpyxl is not installed")
class TestMasterEvidenceXLSX(unittest.TestCase):

    def test_xlsx_has_all_design_doc_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence(
                [_row()], tmp, formats=["xlsx"], warnings=_warnings()
            )["xlsx"]
            self.assertIsNotNone(path)
            wb = openpyxl.load_workbook(path)
            self.assertEqual(
                wb.sheetnames,
                ["Evidence Rows", "Outcome Summary", "Assignment Summary", "Warnings", "README"],
            )

    def test_xlsx_evidence_rows_have_stable_headers_and_semicolon_lists(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence([_row()], tmp, formats=["xlsx"])["xlsx"]
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb["Evidence Rows"]
            headers = [cell.value for cell in ws[1]]
            self.assertEqual(tuple(headers), MASTER_EVIDENCE_EXPORT_COLUMNS)
            course_outcomes_col = headers.index("Course Outcomes") + 1
            self.assertEqual(ws.cell(2, course_outcomes_col).value, "LO1;LO4")
            self.assertEqual(ws.freeze_panes, "A2")

    def test_xlsx_summary_and_warning_sheets_are_populated(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence(
                [_row()], tmp, formats=["xlsx"], warnings=_warnings()
            )["xlsx"]
            wb = openpyxl.load_workbook(path, data_only=True)
            outcome_ws = wb["Outcome Summary"]
            assignment_ws = wb["Assignment Summary"]
            warnings_ws = wb["Warnings"]
            self.assertGreater(outcome_ws.max_row, 1)
            self.assertEqual(assignment_ws.cell(2, 1).value, "PS1 — Asymptotic Analysis")
            self.assertEqual(warnings_ws.cell(2, 1).value, "missing_submission_meta")

    def test_xlsx_readme_records_policy_and_warning_count(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence(
                [_row()], tmp, formats=["xlsx"], warnings=_warnings()
            )["xlsx"]
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb["README"]
            values = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
            self.assertEqual(values["Evidence policy"], "counted_only")
            self.assertIn("1 non-fatal", values["Warnings"])


class TestMasterEvidenceExportAPI(unittest.TestCase):

    def test_requested_formats_return_fixed_paths(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = export_master_evidence([_row()], tmp, formats=["csv", "json"])
            self.assertEqual(set(paths), {"csv", "json"})
            self.assertEqual(Path(paths["csv"]).name, "master_abet_evidence.csv")
            self.assertEqual(Path(paths["json"]).name, "master_abet_evidence.json")

    def test_duplicate_format_names_are_deduplicated(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = export_master_evidence([_row()], tmp, formats=["CSV", "csv", "json"])
            self.assertEqual(set(paths), {"csv", "json"})

    def test_unsupported_format_fails(self):
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaises(ValueError):
                export_master_evidence([_row()], tmp, formats=["pdf"])

    def test_empty_format_list_fails(self):
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaises(ValueError):
                export_master_evidence([_row()], tmp, formats=[])

    def test_explicit_invalid_evidence_policy_fails(self):
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaises(ValueError):
                export_master_evidence(
                    [_row()], tmp, formats=["json"], evidence_policy="bogus"
                )

    def test_mixed_row_policies_are_reported_as_mixed_when_not_overridden(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_master_evidence(
                [_row(), _row(evidence_policy="all", student_id="bob")],
                tmp,
                formats=["json"],
            )["json"]
            payload = json.loads(Path(path).read_text(encoding="utf-8"))
            self.assertEqual(payload["evidence_policy"], "mixed")


if __name__ == "__main__":
    unittest.main()
